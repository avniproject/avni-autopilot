"""Write the Field rules (curated) tab in resources/rules/rules_ai_automation.xlsx.

One row per confirmed semantic cluster. Bodies are normalised per
/curate-rules §6 (drop leading `//SAMPLE…` marker, repair `use strict';`
→ `"use strict";`); no semantic edits.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

WORKBOOK = Path("resources/rules/rules_ai_automation.xlsx")
RAW_SHEET = "Field rules"
CURATED_SHEET = "Field rules (curated)"

# Cluster reps (row in raw sheet) + signed-off intent prompt.
CLUSTERS: list[tuple[int, str]] = [
    (3544,
     "show this field only when an earlier question's selected option was a specific one"),
    (9,
     "show this field only when an earlier question's answer matches a specific option "
     "(e.g. show details when 'Consent given' is Yes)"),
    (157,
     "show this field only when the subject's age is above or below a threshold "
     "(e.g. show child-only details when the subject is under 18)"),
    (12,
     "show this field only when a numeric value recorded earlier crosses a threshold "
     "(e.g. show capacity when 'Number of trolleys' is more than 0)"),
    (22,
     "show this field only when several earlier answers together satisfy a combined "
     "condition (e.g. weight under 2.5 kg AND KMC was done)"),
    (8940,
     "show this field only when an earlier observation has been filled in, regardless "
     "of its value"),
    (1192,
     "show this field only when the cancellation form recorded a specific reason "
     "(e.g. show death details when cancellation reason is Child Death)"),
    (4221,
     "show this field only until specific values have been recorded in a prior "
     "encounter (e.g. show TD Booster until both TD 1 and TD 2 have already been "
     "given)"),
    (12132,
     "show this field only when a certain number of days have passed since a "
     "previously recorded date (e.g. after 126 days since LMP)"),
    (9734,
     "pre-fill this field with the value of another observation already recorded "
     "for the subject (e.g. copy phone number from registration)"),
    (3294,
     "compute this field's value from earlier observations "
     "(e.g. total = sum of male + female members)"),
    (4864,
     "block save when the date entered in this field is in the past or future "
     "compared to today"),
    (15,
     "block save when this field's value is inconsistent with a related earlier "
     "field (e.g. machine end time must be later than start time)"),
    (307,
     "hide specific coded answer options for this field based on an earlier answer "
     "(e.g. hide C-section and Assisted delivery when place of delivery is at home)"),
]

CURATED_HEADERS = ["org_name", "form_name", "field_name", "rule", "prompt"]

SAMPLE_MARKER_RE = re.compile(
    r"^\s*//\s*SAMPLE\s+(RULE\s+EXAMPLE|EDIT\s+FORM\s+RULE)\s*$\n?",
    re.IGNORECASE | re.MULTILINE,
)
USE_STRICT_BUG_RE = re.compile(r"^[\t ]*use strict';", re.MULTILINE)


def normalise(body: str) -> tuple[str, int, int]:
    new, dropped = SAMPLE_MARKER_RE.subn("", body, count=1)
    new, repaired = USE_STRICT_BUG_RE.subn('"use strict";', new)
    return new, dropped, repaired


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    raw = wb[RAW_SHEET]
    cur = wb[CURATED_SHEET]

    if cur.max_row > 1 or (cur.max_row == 1 and any(c.value for c in cur[1])):
        raise SystemExit(
            f"{CURATED_SHEET!r} is not empty (max_row={cur.max_row}). "
            "Refusing to clobber — clear it or use a numbered variant."
        )

    raw_headers = {c.value: i + 1 for i, c in enumerate(raw[1])}
    for k in ("org_name", "form_name", "field_name", "rule"):
        if k not in raw_headers:
            raise SystemExit(f"Raw sheet missing column {k!r}: {list(raw_headers)}")

    bold = Font(bold=True)
    for col, name in enumerate(CURATED_HEADERS, start=1):
        cur.cell(1, col, name).font = bold

    cur.column_dimensions["A"].width = 22
    cur.column_dimensions["B"].width = 32
    cur.column_dimensions["C"].width = 30
    cur.column_dimensions["D"].width = 90
    cur.column_dimensions["E"].width = 80

    total_dropped = 0
    total_repaired = 0
    written = 0

    for cluster_idx, (rep_row, prompt) in enumerate(CLUSTERS, start=1):
        org   = raw.cell(rep_row, raw_headers["org_name"]).value or ""
        form  = raw.cell(rep_row, raw_headers["form_name"]).value or ""
        field = raw.cell(rep_row, raw_headers["field_name"]).value or ""
        rule  = raw.cell(rep_row, raw_headers["rule"]).value or ""
        if not isinstance(rule, str) or not rule.strip():
            raise SystemExit(f"Representative row {rep_row} has no rule body.")
        clean, dropped, repaired = normalise(rule)
        total_dropped += dropped
        total_repaired += repaired

        out_row = cluster_idx + 1
        cur.cell(out_row, 1, org)
        cur.cell(out_row, 2, form)
        cur.cell(out_row, 3, field)
        rule_cell = cur.cell(out_row, 4, clean)
        rule_cell.alignment = Alignment(wrap_text=True, vertical="top")
        prompt_cell = cur.cell(out_row, 5, prompt)
        prompt_cell.alignment = Alignment(wrap_text=True, vertical="top")
        written += 1

    wb.save(WORKBOOK)
    print(f"Wrote {written} rows to {CURATED_SHEET!r}.")
    print(f"Normalisation: //SAMPLE dropped={total_dropped}, "
          f"use strict' repaired={total_repaired}")


if __name__ == "__main__":
    main()
