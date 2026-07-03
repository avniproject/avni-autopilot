"""Cluster Field / Page rules in resources/rules/rules_ai_automation.xlsx by
structural fingerprint. Outputs a JSON manifest of clusters with one
representative per cluster — fed back to Claude for naming, prompt-writing,
and curated-tab population.

The fingerprint is intentionally cheap and deterministic. It captures the
shape of the rule, not its identifiers, so two orgs writing the same
idiom against different field names cluster together.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

WORKBOOK = Path("resources/rules/rules_ai_automation.xlsx")


# ── Tokenisation / signature ────────────────────────────────────────────────

# Tokens that name a structural feature of a rule body. Order matters only
# for readability of the resulting signature string.
SIGNATURE_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("ENTITY_individual",       re.compile(r"params\.entity[^;\n]*=\s*params\.entity[^;\n]*$", re.M)),  # rarely used
    ("ENTITY_programEnrolment", re.compile(r"\bprogramEnrolment\s*=\s*params\.entity\b")),
    ("ENTITY_programEncounter", re.compile(r"\bprogramEncounter\s*=\s*params\.entity\b")),
    ("ENTITY_encounter",        re.compile(r"\bencounter\s*=\s*params\.entity\b(?!\.programEnrolment)")),
    ("ENTITY_individual_alt",   re.compile(r"\bindividual\s*=\s*params\.entity\b")),

    ("USE_formElement",      re.compile(r"\bparams\.formElement\b")),
    ("USE_formElementGroup", re.compile(r"\bparams\.formElementGroup\b")),

    ("RET_visibility",       re.compile(r"\b(visibility\s*[:=])|FormElementStatus\(.*?,\s*(true|false)")),
    ("RET_value",            re.compile(r"FormElementStatus\([^,]+,[^,]+,\s*[^,)]+")),
    ("RET_validationErrors", re.compile(r"validationErrors|complicationsBuilder")),
    ("RET_mandatory",        re.compile(r"\bmandatory\s*[:=]\s*(true|false)")),
    ("RET_answerFilter",     re.compile(r"answersToShow|answersToHide|answerFilter|skipAnswers")),

    ("HELPER_getObsReadable",   re.compile(r"\bgetObservationReadableValue\b")),
    ("HELPER_getObs",           re.compile(r"\bgetObservation\(")),
    ("HELPER_getObsValue",      re.compile(r"\bgetObservationValue\b")),
    ("HELPER_findCancelObs",    re.compile(r"\bfindCancelEncounterObservationReadableValue\b|\bgetCancelObservation\b")),
    ("HELPER_RuleCondition",    re.compile(r"\bRuleCondition\b")),
    ("HELPER_encounters",       re.compile(r"\.getEncounters\(|\.nonVoidedEncounters\b|\.scheduledEncounters\b")),
    ("HELPER_lastEnc",          re.compile(r"\blastFulfilledEncounter\b|\blastEncounter\b")),
    ("HELPER_individualAge",    re.compile(r"\.getAgeInYears\(|\.getAgeInMonths\(|moment\(.*dateOfBirth")),
    ("HELPER_uniqueByType",     re.compile(r"\buniqueByType\b")),
    ("HELPER_moment",           re.compile(r"\bmoment\(")),
    ("HELPER_complications",    re.compile(r"\bcomplicationsBuilder\b|\bComplicationsBuilder\b")),

    ("COND_yes_no",             re.compile(r"=== ?\"Yes\"|=== ?'Yes'|=== ?\"No\"|=== ?'No'", re.I)),
    ("COND_dateOrdering",       re.compile(r"\bisBefore\(|\bisAfter\(")),
    ("COND_numericRange",       re.compile(r"<=?\s*\d|>=?\s*\d")),

    ("STR_arrowFn",             re.compile(r"\(\{\s*params\s*,\s*imports\s*\}\)\s*=>")),
    ("STR_voidedFilter",        re.compile(r"!?voided\b|\bvoided ===")),
]


# Helpers for fingerprint cleanup
SAMPLE_MARKER_RE = re.compile(
    r"^\s*//\s*SAMPLE\s+(RULE\s+EXAMPLE|EDIT\s+FORM\s+RULE)\s*$\n?",
    re.IGNORECASE | re.MULTILINE,
)
USE_STRICT_BUG_RE = re.compile(r"^[\t ]*use strict';", re.MULTILINE)


def normalise_rule_body(body: str) -> tuple[str, dict[str, int]]:
    """Apply the two-allowed cleanups; return cleaned body + per-fix counts."""
    counts = {"sample_marker_dropped": 0, "use_strict_repaired": 0}
    new, n = SAMPLE_MARKER_RE.subn("", body, count=1)
    counts["sample_marker_dropped"] = n
    new, n = USE_STRICT_BUG_RE.subn('"use strict";', new)
    counts["use_strict_repaired"] = n
    return new, counts


def fingerprint(body: str) -> str:
    """Deterministic structural signature of a rule body."""
    if not body:
        return "EMPTY"
    hits: list[str] = []
    for tag, pat in SIGNATURE_PATTERNS:
        if pat.search(body):
            hits.append(tag)
    # Add a coarse size bucket so radically different bodies with the same
    # helpers don't collapse together.
    n_lines = body.count("\n") + 1
    if   n_lines <= 5:   bucket = "size_xs"
    elif n_lines <= 15:  bucket = "size_s"
    elif n_lines <= 40:  bucket = "size_m"
    elif n_lines <= 100: bucket = "size_l"
    else:                bucket = "size_xl"
    hits.append(bucket)
    return "|".join(hits) if hits else "EMPTY_SIGNATURE"


# ── Cluster build ───────────────────────────────────────────────────────────


def cluster_sheet(sheet_name: str, label_col: str, out_path: Path) -> None:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[sheet_name]

    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    required = ["org_name", "form_name", label_col, "rule"]
    for k in required:
        if k not in idx:
            raise SystemExit(f"sheet {sheet_name!r} missing column {k!r} — got {header}")

    clusters: dict[str, list[dict]] = defaultdict(list)
    norm_counts = Counter()
    total_rows = 0
    empty_rule_rows = 0

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c + 1).value for c in range(len(header))]
        rule = row[idx["rule"]]
        if not (isinstance(rule, str) and rule.strip()):
            empty_rule_rows += 1
            continue
        total_rows += 1
        clean_rule, fix_counts = normalise_rule_body(rule)
        for k, v in fix_counts.items():
            norm_counts[k] += v
        sig = fingerprint(clean_rule)
        clusters[sig].append({
            "row": r,
            "org_name": row[idx["org_name"]] or "",
            "form_name": row[idx["form_name"]] or "",
            label_col: row[idx[label_col]] or "",
            "rule_raw": rule,
            "rule_clean": clean_rule,
            "line_count": clean_rule.count("\n") + 1,
        })

    # Order clusters by size desc.
    ordered = sorted(clusters.items(), key=lambda kv: -len(kv[1]))
    cluster_records = []
    for cid, members in ordered:
        # Representative = shortest clean body in the cluster.
        rep = min(members, key=lambda m: m["line_count"])
        cluster_records.append({
            "signature": cid,
            "size": len(members),
            "representative": {
                "row": rep["row"],
                "org_name": rep["org_name"],
                "form_name": rep["form_name"],
                label_col: rep[label_col],
                "line_count": rep["line_count"],
                "rule_clean": rep["rule_clean"],
            },
            "sample_members": [
                {
                    "row": m["row"],
                    "org_name": m["org_name"],
                    "form_name": m["form_name"],
                    label_col: m[label_col],
                }
                for m in members[:5]
            ],
        })

    out = {
        "sheet": sheet_name,
        "label_col": label_col,
        "stats": {
            "rows_seen": total_rows,
            "rows_skipped_empty_rule": empty_rule_rows,
            "clusters_found": len(cluster_records),
            "normalisation": dict(norm_counts),
        },
        "clusters": cluster_records,
    }
    out_path.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print(f"{sheet_name}: rows={total_rows} clusters={len(cluster_records)} → {out_path}")


# ── CLI ─────────────────────────────────────────────────────────────────────


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--out-dir", default="scripts/_cluster_out")
    args = p.parse_args()
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    cluster_sheet("Page rules",  "page_name",  out_dir / "page_rules_clusters.json")
    cluster_sheet("Field rules", "field_name", out_dir / "field_rules_clusters.json")
