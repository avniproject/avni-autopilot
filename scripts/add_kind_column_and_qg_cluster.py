"""Two one-off chores on resources/rules/rules_ai_automation.xlsx:

1. Append cluster #15 to `Field rules (curated)` — the
   `questionGroupValueInEncounter` idiom (visibility inside a repeating
   group). Representative: r2358 of the raw `Field rules` sheet.

2. Add a `kind` column to both `(curated)` tabs so future ingest can
   filter by behaviour. Mapping per /curate-rules sign-off:

   Field rules (curated):
     rows 2-10  → visibility   (clusters 1-9)
     row  11    → value        (cluster 10 — pre-fill)
     row  12    → value        (cluster 11 — compute)
     row  13    → validation   (cluster 12 — date)
     row  14    → validation   (cluster 13 — cross-field)
     row  15    → answerFilter (cluster 14)
     row  16    → visibility   (new cluster 15 — questionGroupValueIn*)

   Page rules (curated):
     rows 2-9   → visibility   (all 8 clusters)

Idempotent: if the `kind` column already exists, this script overwrites
it with the canonical mapping rather than appending a duplicate.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

WORKBOOK = Path("resources/rules/rules_ai_automation.xlsx")

FIELD_CLUSTER_15_PROMPT = (
    "show this field inside a repeating section only when another value "
    "within the same row meets a numeric condition (e.g. show 'Marks for Q11' "
    "only when 'Total questions' is between 11 and 20)"
)

FIELD_KIND_BY_ROW = {
    2:  "visibility",
    3:  "visibility",
    4:  "visibility",
    5:  "visibility",
    6:  "visibility",
    7:  "visibility",
    8:  "visibility",
    9:  "visibility",
    10: "visibility",
    11: "value",
    12: "value",
    13: "validation",
    14: "validation",
    15: "answerFilter",
    16: "visibility",
}

SAMPLE_MARKER_RE = re.compile(
    r"^\s*//\s*SAMPLE\s+(RULE\s+EXAMPLE|EDIT\s+FORM\s+RULE)\s*$\n?",
    re.IGNORECASE | re.MULTILINE,
)
USE_STRICT_BUG_RE = re.compile(r"^[\t ]*use strict';", re.MULTILINE)


def normalise(body: str) -> tuple[str, int, int]:
    new, dropped = SAMPLE_MARKER_RE.subn("", body, count=1)
    new, repaired = USE_STRICT_BUG_RE.subn('"use strict";', new)
    return new, dropped, repaired


def append_qg_cluster(wb: openpyxl.Workbook) -> None:
    raw = wb["Field rules"]
    cur = wb["Field rules (curated)"]

    raw_headers = {c.value: i + 1 for i, c in enumerate(raw[1])}
    rep_row = 2358
    rule = raw.cell(rep_row, raw_headers["rule"]).value
    if not isinstance(rule, str) or not rule.strip():
        raise SystemExit(f"r{rep_row} has no rule body in 'Field rules'.")
    clean, dropped, repaired = normalise(rule)

    # Find next free row in curated tab (header + 14 existing = max_row 15).
    out_row = cur.max_row + 1
    if out_row != 16:
        raise SystemExit(
            f"Expected curated tab to have 15 rows before append, found {cur.max_row}. "
            "Refusing to clobber — verify state."
        )
    cur.cell(out_row, 1, raw.cell(rep_row, raw_headers["org_name"]).value or "")
    cur.cell(out_row, 2, raw.cell(rep_row, raw_headers["form_name"]).value or "")
    cur.cell(out_row, 3, raw.cell(rep_row, raw_headers["field_name"]).value or "")
    rule_cell = cur.cell(out_row, 4, clean)
    rule_cell.alignment = Alignment(wrap_text=True, vertical="top")
    prompt_cell = cur.cell(out_row, 5, FIELD_CLUSTER_15_PROMPT)
    prompt_cell.alignment = Alignment(wrap_text=True, vertical="top")
    print(f"Appended cluster 15 at row {out_row}. Normalisation: "
          f"//SAMPLE dropped={dropped}, use strict' repaired={repaired}")


def add_kind_column(wb: openpyxl.Workbook, sheet: str, mapping: dict[int, str]) -> None:
    ws = wb[sheet]
    # Find or create the `kind` column.
    headers = [c.value for c in ws[1]]
    bold = Font(bold=True)
    if "kind" in headers:
        col = headers.index("kind") + 1
    else:
        col = ws.max_column + 1
        ws.cell(1, col, "kind").font = bold
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 14
    for row, kind in mapping.items():
        ws.cell(row, col, kind)
    print(f"{sheet}: kind column at col {col}, populated {len(mapping)} rows.")


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    append_qg_cluster(wb)
    add_kind_column(wb, "Field rules (curated)", FIELD_KIND_BY_ROW)
    add_kind_column(wb, "Page rules (curated)", {r: "visibility" for r in range(2, 10)})
    wb.save(WORKBOOK)
    print(f"Saved {WORKBOOK}")


if __name__ == "__main__":
    main()
