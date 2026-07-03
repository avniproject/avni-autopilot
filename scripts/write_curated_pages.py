"""Write the Page rules (curated) tab in resources/rules/rules_ai_automation.xlsx.

One row per confirmed cluster. The body is the cluster's representative,
normalised per /curate-rules §6 (drop leading `//SAMPLE…` marker, repair
`use strict';` → `"use strict";`). No semantic edits.

Reps + prompts are hard-coded here (the cluster confirmation step is human
sign-off, not derivable). Run once; idempotent if the curated tab is empty.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

WORKBOOK = Path("resources/rules/rules_ai_automation.xlsx")
RAW_SHEET = "Page rules"
CURATED_SHEET = "Page rules (curated)"

# Cluster reps (row in raw sheet) + the intent prompt the user signed off on.
CLUSTERS: list[tuple[int, str]] = [
    (213,
     "show this page only when an earlier question's selected option was a specific one"),
    (29,
     "show this page only when a previously recorded value equals a specific string "
     "(e.g. show 'Abortion Details' only when 'Outcome of pregnancy' is Abortion)"),
    (2,
     "show this page only when a numeric value recorded earlier is at or above a "
     "threshold (e.g. show referral page when 'Number of places visited' is 7 or more)"),
    (60,
     "show this page only when several earlier answers together satisfy a combined "
     "condition (e.g. period of death is pregnancy AND pregnancy weeks >= 6 AND "
     "referred is Yes)"),
    (3,
     "show this page only when an earlier question's answer matches a specific option "
     "(e.g. show details when 'Consent given' is Yes)"),
    (131,
     "show this page only when the subject's age is below a threshold "
     "(e.g. show child-specific details when the subject is under 18)"),
    (1020,
     "show this page only during a specific month or season of the year "
     "(e.g. show inputs page only in June)"),
    (74,
     "show this page only when a time-based condition involving the current encounter "
     "and prior encounters is met"),
]

CURATED_HEADERS = ["org_name", "form_name", "page_name", "rule", "prompt"]

SAMPLE_MARKER_RE = re.compile(
    r"^\s*//\s*SAMPLE\s+(RULE\s+EXAMPLE|EDIT\s+FORM\s+RULE)\s*$\n?",
    re.IGNORECASE | re.MULTILINE,
)
USE_STRICT_BUG_RE = re.compile(r"^[\t ]*use strict';", re.MULTILINE)


def normalise(body: str) -> tuple[str, int, int]:
    new, dropped = SAMPLE_MARKER_RE.subn("", body, count=1)
    new, repaired = USE_STRICT_BUG_RE.subn('"use strict";', new)
    return new, dropped, repaired


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    raw = wb[RAW_SHEET]
    cur = wb[CURATED_SHEET]

    # Sanity: refuse to overwrite a populated curated tab.
    if cur.max_row > 1 or (cur.max_row == 1 and any(c.value for c in cur[1])):
        raise SystemExit(
            f"{CURATED_SHEET!r} is not empty (max_row={cur.max_row}). "
            "Refusing to clobber — clear it or use a numbered variant."
        )

    raw_headers = {c.value: i + 1 for i, c in enumerate(raw[1])}
    for k in ("org_name", "form_name", "page_name", "rule"):
        if k not in raw_headers:
            raise SystemExit(f"Raw sheet missing column {k!r}: {list(raw_headers)}")

    # Header row.
    bold = Font(bold=True)
    for col, name in enumerate(CURATED_HEADERS, start=1):
        cell = cur.cell(1, col, name)
        cell.font = bold

    # Wider columns + wrapped text so the JS body is legible.
    cur.column_dimensions["A"].width = 22  # org_name
    cur.column_dimensions["B"].width = 32  # form_name
    cur.column_dimensions["C"].width = 28  # page_name
    cur.column_dimensions["D"].width = 90  # rule
    cur.column_dimensions["E"].width = 80  # prompt

    total_dropped = 0
    total_repaired = 0
    written = 0

    for cluster_idx, (rep_row, prompt) in enumerate(CLUSTERS, start=1):
        org   = raw.cell(rep_row, raw_headers["org_name"]).value or ""
        form  = raw.cell(rep_row, raw_headers["form_name"]).value or ""
        page  = raw.cell(rep_row, raw_headers["page_name"]).value or ""
        rule  = raw.cell(rep_row, raw_headers["rule"]).value or ""
        if not isinstance(rule, str) or not rule.strip():
            raise SystemExit(f"Representative row {rep_row} has no rule body.")
        clean, dropped, repaired = normalise(rule)
        total_dropped += dropped
        total_repaired += repaired

        out_row = cluster_idx + 1
        cur.cell(out_row, 1, org)
        cur.cell(out_row, 2, form)
        cur.cell(out_row, 3, page)
        rule_cell = cur.cell(out_row, 4, clean)
        rule_cell.alignment = Alignment(wrap_text=True, vertical="top")
        prompt_cell = cur.cell(out_row, 5, prompt)
        prompt_cell.alignment = Alignment(wrap_text=True, vertical="top")
        written += 1

    wb.save(WORKBOOK)
    print(f"Wrote {written} rows to {CURATED_SHEET!r}.")
    print(f"Normalisation: //SAMPLE dropped={total_dropped}, "
          f"use strict' repaired={total_repaired}")


if __name__ == "__main__":
    main()
