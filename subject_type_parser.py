#!/usr/bin/env python3
"""
LangGraph pipeline: parse Subject Types from an Avni modelling Excel sheet
and produce subjectTypes.json + operationalSubjectTypes.json.

Usage:
    python subject_type_parser.py \
        --input  resources/input/"SRIJAN Modelling - CONSISTENT (1).xlsx" \
        --output resources/output/Srijan
"""

from __future__ import annotations

import argparse
import json
import os
import uuid
from typing import TypedDict

import openpyxl
from langgraph.graph import END, StateGraph

# ── State ────────────────────────────────────────────────────────────────────


class PipelineState(TypedDict):
    file_path: str
    output_dir: str
    raw_rows: list[dict]           # dicts keyed by column header
    subject_types: list[dict]      # normalised subject type records
    subject_types_json: list[dict]
    operational_subject_types_json: dict
    errors: list[str]


# ── Type → field defaults ────────────────────────────────────────────────────

_TYPE_DEFAULTS: dict[str, dict] = {
    "Person": {
        "type": "Person",
        "allowEmptyLocation": False,
        "lastNameOptional": False,
        "allowProfilePicture": False,
        "shouldSyncByLocation": True,
        "household": False,
        "group": False,
        "directlyAssignable": False,
    },
    "Individual": {
        "type": "Individual",
        "allowEmptyLocation": False,
        "lastNameOptional": False,
        "allowProfilePicture": False,
        "shouldSyncByLocation": True,
        "household": False,
        "group": False,
        "directlyAssignable": False,
    },
    "Household": {
        "type": "Household",
        "allowEmptyLocation": False,
        "lastNameOptional": False,
        "allowProfilePicture": False,
        "shouldSyncByLocation": True,
        "household": True,
        "group": True,
        "directlyAssignable": False,
    },
    "User": {
        "type": "Individual",
        "allowEmptyLocation": True,
        "lastNameOptional": False,
        "allowProfilePicture": False,
        "shouldSyncByLocation": False,
        "household": False,
        "group": False,
        "directlyAssignable": False,
    },
}

_SETTINGS_DEFAULT = {
    "displayRegistrationDetails": True,
    "displayPlannedEncounters": True,
}


# ── Node: read_sheet ─────────────────────────────────────────────────────────


def read_sheet(state: PipelineState) -> PipelineState:
    """Open the Excel file and extract all non-empty rows from 'Subject Types' sheet."""
    file_path = state["file_path"]
    errors = list(state.get("errors", []))

    if not os.path.exists(file_path):
        errors.append(f"File not found: {file_path}")
        return {**state, "raw_rows": [], "errors": errors}

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        errors.append(f"Cannot open workbook: {exc}")
        return {**state, "raw_rows": [], "errors": errors}

    sheet_name = next(
        (s for s in wb.sheetnames if s.strip().lower() == "subject types"), None
    )
    if sheet_name is None:
        errors.append(f"No 'Subject Types' sheet found. Sheets: {wb.sheetnames}")
        return {**state, "raw_rows": [], "errors": errors}

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    # First non-empty row is the header
    headers: list[str] = []
    for row in rows_iter:
        if any(c is not None for c in row):
            headers = [str(c).strip() if c is not None else "" for c in row]
            break

    raw_rows: list[dict] = []
    for row in rows_iter:
        if not any(c is not None for c in row):
            continue
        record = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        # Skip completely blank rows (all values None / empty string)
        if all(v is None or str(v).strip() == "" for v in record.values()):
            continue
        raw_rows.append(record)

    return {**state, "raw_rows": raw_rows, "errors": errors}


# ── Node: parse_subject_types ─────────────────────────────────────────────────


def parse_subject_types(state: PipelineState) -> PipelineState:
    """Normalise raw rows into subject type records with UUIDs and type flags."""
    raw_rows = state.get("raw_rows", [])
    errors = list(state.get("errors", []))
    subject_types: list[dict] = []

    # Flexible header matching
    def _find(row: dict, *candidates: str):
        for key in row:
            if key.strip().lower() in [c.lower() for c in candidates]:
                val = row[key]
                return str(val).strip() if val is not None else ""
        return ""

    for row in raw_rows:
        name = _find(row, "Subject Type Name", "Name")
        raw_type = _find(row, "Type")

        if not name:
            continue

        # Normalise type: match case-insensitively, fall back to "Individual"
        matched_type = next(
            (t for t in _TYPE_DEFAULTS if t.lower() == raw_type.lower()),
            "Individual",
        )
        if matched_type != raw_type and raw_type:
            errors.append(
                f"Unknown type '{raw_type}' for '{name}'; defaulted to 'Individual'."
            )

        subject_types.append(
            {
                "name": name,
                "uuid": str(uuid.uuid4()),
                "raw_type": matched_type,
            }
        )

    return {**state, "subject_types": subject_types, "errors": errors}


# ── Node: generate_json ───────────────────────────────────────────────────────


def generate_json(state: PipelineState) -> PipelineState:
    """Build subjectTypes.json list and operationalSubjectTypes.json dict."""
    subject_types = state.get("subject_types", [])

    subject_types_json: list[dict] = []
    operational_entries: list[dict] = []

    for st in subject_types:
        defaults = _TYPE_DEFAULTS[st["raw_type"]]
        st_entry = {
            "name": st["name"],
            "uuid": st["uuid"],
            "active": True,
            "type": defaults["type"],
            "subjectSummaryRule": "",
            "programEligibilityCheckRule": "",
            "memberAdditionEligibilityCheckRule": "",
            "allowEmptyLocation": defaults["allowEmptyLocation"],
            "allowMiddleName": False,
            "lastNameOptional": defaults["lastNameOptional"],
            "allowProfilePicture": defaults["allowProfilePicture"],
            "uniqueName": False,
            "shouldSyncByLocation": defaults["shouldSyncByLocation"],
            "settings": dict(_SETTINGS_DEFAULT),
            "household": defaults["household"],
            "group": defaults["group"],
            "directlyAssignable": defaults["directlyAssignable"],
            "voided": False,
        }
        subject_types_json.append(st_entry)

        operational_entries.append(
            {
                "uuid": str(uuid.uuid4()),
                "subjectType": {"uuid": st["uuid"], "voided": False},
                "name": st["name"],
                "voided": False,
            }
        )

    operational_subject_types_json = {"operationalSubjectTypes": operational_entries}

    return {
        **state,
        "subject_types_json": subject_types_json,
        "operational_subject_types_json": operational_subject_types_json,
    }


# ── Node: save_files ──────────────────────────────────────────────────────────


def save_files(state: PipelineState) -> PipelineState:
    """Write subjectTypes.json and operationalSubjectTypes.json to output_dir."""
    output_dir = state["output_dir"]
    errors = list(state.get("errors", []))

    os.makedirs(output_dir, exist_ok=True)

    st_path = os.path.join(output_dir, "subjectTypes.json")
    ost_path = os.path.join(output_dir, "operationalSubjectTypes.json")

    try:
        with open(st_path, "w", encoding="utf-8") as f:
            json.dump(state["subject_types_json"], f, indent=2, ensure_ascii=False)
        with open(ost_path, "w", encoding="utf-8") as f:
            json.dump(
                state["operational_subject_types_json"], f, indent=2, ensure_ascii=False
            )
    except OSError as exc:
        errors.append(f"Failed to write output files: {exc}")

    return {**state, "errors": errors}


# ── Conditional edge: abort if errors before saving ──────────────────────────


def _should_save(state: PipelineState) -> str:
    """Skip save_files if any hard errors were recorded earlier in the pipeline."""
    if state.get("errors") and not state.get("subject_types"):
        return "abort"
    return "save"


# ── Graph assembly ────────────────────────────────────────────────────────────


def build_graph() -> StateGraph:
    graph = StateGraph(PipelineState)

    graph.add_node("read_sheet", read_sheet)
    graph.add_node("parse_subject_types", parse_subject_types)
    graph.add_node("generate_json", generate_json)
    graph.add_node("save_files", save_files)

    graph.set_entry_point("read_sheet")
    graph.add_edge("read_sheet", "parse_subject_types")
    graph.add_edge("parse_subject_types", "generate_json")
    graph.add_conditional_edges(
        "generate_json",
        _should_save,
        {"save": "save_files", "abort": END},
    )
    graph.add_edge("save_files", END)

    return graph.compile()


# ── Entry point ───────────────────────────────────────────────────────────────


def run(file_path: str, output_dir: str) -> PipelineState:
    pipeline = build_graph()
    initial: PipelineState = {
        "file_path": file_path,
        "output_dir": output_dir,
        "raw_rows": [],
        "subject_types": [],
        "subject_types_json": [],
        "operational_subject_types_json": {},
        "errors": [],
    }
    return pipeline.invoke(initial)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Parse Subject Types from Avni modelling Excel")
    parser.add_argument("--input", required=True, help="Path to the input .xlsx file")
    parser.add_argument("--output", required=True, help="Directory to write JSON output files")
    args = parser.parse_args()

    result = run(args.input, args.output)

    if result["errors"]:
        print("Warnings / errors:")
        for e in result["errors"]:
            print(f"  • {e}")

    count = len(result["subject_types_json"])
    print(f"Parsed {count} subject type(s) → {args.output}/subjectTypes.json + operationalSubjectTypes.json")
