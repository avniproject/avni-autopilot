"""CLI for maintaining the rule knowledge base.

Day-to-day commands (porcelain):

  helpers           Refresh entity helpers end-to-end (sync + enrich-use-when
                    + rebuild). Use after pulling new avni-models source or
                    hand-editing helper files. Pass --skip-enrich to avoid
                    the Haiku cost.
  examples          Refresh one rule-kind's example corpus (ingest-examples
                    + rebuild). Use after editing one curated tab.
  examples-all      Ingest every wired RuleKind's curated tab from
                    `resources/rules/rules_ai_automation.xlsx` and rebuild
                    once. Use for first-time setup or after bulk edits
                    across tabs.

Surgical commands (plumbing, run individually when you know what you need):

  sync              Walk avni-models source and (re)generate per-entity helper
                    catalogs under resources/rules/helpers/entities/. Existing
                    `signature` / `use_when` / `example_snippet` fields are
                    preserved on a merge by method name.
  enrich-use-when   Read each method's body from the avni-models source and ask
                    Claude Haiku to write a one-sentence `use_when`. Skips any
                    method whose `use_when` is already non-empty.
  ingest-examples   Read the curated rules tab from the source xlsx and write
                    one Markdown file per row under
                    resources/rules/examples/<rule_kind>/.
  rebuild           Force a full re-embed of the helper + example catalogs into
                    resources/rules/.embeddings.json (Voyage).

Invoke either via the project script (`avni-rules-kb <cmd> ...`) or directly
(`python -m domain.rules.kb_cli <cmd> ...`). See
specs/VISIT_SCHEDULE_RULE_SDD.md §6.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from collections.abc import Iterator
from pydantic import BaseModel, Field

import config  # noqa: F401  — imported for the .env-loading side effect
from domain.rules.accessors import concept_accessor_call_regex
from domain.rules.knowledge_base import KnowledgeBase
from domain.rules.rule_spec import RuleKind

log = logging.getLogger("rules.kb")


# ── Defaults ──────────────────────────────────────────────────────────────────

_REPO_ROOT = Path(__file__).resolve().parents[3]
_AVNI_MODELS_DEFAULT = _REPO_ROOT.parent / "avni-models" / "src"
_XLSX_DEFAULT = _REPO_ROOT / "resources" / "rules" / "rules_ai_automation.xlsx"
_CURATED_TAB_DEFAULT = "VS rule (curated)"
_RULES_ROOT = _REPO_ROOT / "resources" / "rules"


# ── Entry point ───────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    parser = argparse.ArgumentParser(prog="avni-rules-kb", description=__doc__)
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_sync = sub.add_parser("sync", help="Sync per-entity helpers from avni-models source.")
    p_sync.add_argument("--source", type=Path, default=_AVNI_MODELS_DEFAULT)
    p_sync.add_argument("--root", type=Path, default=_RULES_ROOT)

    p_ingest = sub.add_parser("ingest-examples",
                              help="Generate example markdown from the curated xlsx tab.")
    p_ingest.add_argument("--xlsx", type=Path, default=_XLSX_DEFAULT)
    p_ingest.add_argument("--tab", type=str, default=_CURATED_TAB_DEFAULT)
    p_ingest.add_argument("--root", type=Path, default=_RULES_ROOT)
    p_ingest.add_argument("--rule-kind", type=str, default=RuleKind.VISIT_SCHEDULE.value,
                          help="Subdirectory under examples/ and frontmatter value.")

    p_rebuild = sub.add_parser("rebuild", help="Re-embed the catalog into the cache.")
    p_rebuild.add_argument("--root", type=Path, default=_RULES_ROOT)

    p_enrich = sub.add_parser(
        "enrich-use-when",
        help="Populate empty use_when fields by asking Claude Haiku to read each method's body.",
    )
    p_enrich.add_argument("--source", type=Path, default=_AVNI_MODELS_DEFAULT)
    p_enrich.add_argument("--root", type=Path, default=_RULES_ROOT)
    p_enrich.add_argument("--batch-size", type=int, default=20,
                          help="How many methods to annotate per Haiku call.")
    p_enrich.add_argument("--dry-run", action="store_true",
                          help="Print proposed annotations without writing.")

    # ── Composite porcelain commands ─────────────────────────────────────────
    p_helpers = sub.add_parser(
        "helpers",
        help="Refresh entity helpers end-to-end (sync + enrich-use-when + rebuild).",
    )
    p_helpers.add_argument("--source", type=Path, default=_AVNI_MODELS_DEFAULT)
    p_helpers.add_argument("--root", type=Path, default=_RULES_ROOT)
    p_helpers.add_argument("--batch-size", type=int, default=20,
                           help="Methods per Haiku call during enrich.")
    p_helpers.add_argument("--skip-enrich", action="store_true",
                           help="Skip Haiku annotation (saves API cost on a quick sync).")

    p_examples = sub.add_parser(
        "examples",
        help="Refresh example corpus end-to-end (ingest-examples + rebuild).",
    )
    p_examples.add_argument("--xlsx", type=Path, default=_XLSX_DEFAULT)
    p_examples.add_argument("--tab", type=str, default=_CURATED_TAB_DEFAULT)
    p_examples.add_argument("--root", type=Path, default=_RULES_ROOT)
    p_examples.add_argument("--rule-kind", type=str,
                            default=RuleKind.VISIT_SCHEDULE.value,
                            help="Subdirectory under examples/ and frontmatter value.")

    p_examples_all = sub.add_parser(
        "examples-all",
        help="Ingest curated tabs for every wired RuleKind and rebuild once.",
    )
    p_examples_all.add_argument("--xlsx", type=Path, default=_XLSX_DEFAULT)
    p_examples_all.add_argument("--root", type=Path, default=_RULES_ROOT)

    args = parser.parse_args(argv)

    if args.cmd == "sync":
        return cmd_sync(args.source, args.root)
    if args.cmd == "ingest-examples":
        return cmd_ingest_examples(args.xlsx, args.tab, args.root, args.rule_kind)
    if args.cmd == "rebuild":
        return cmd_rebuild(args.root)
    if args.cmd == "enrich-use-when":
        return cmd_enrich_use_when(args.source, args.root, args.batch_size, args.dry_run)
    if args.cmd == "helpers":
        return cmd_helpers(args.source, args.root, args.batch_size, args.skip_enrich)
    if args.cmd == "examples":
        return cmd_examples(args.xlsx, args.tab, args.root, args.rule_kind)
    if args.cmd == "examples-all":
        return cmd_examples_all(args.xlsx, args.root)
    parser.error(f"unknown command: {args.cmd!r}")
    return 2


# Each RuleKind's curated tab inside
# `resources/rules/rules_ai_automation.xlsx`. See FORM_LEVEL_RULES_SDD §7
# and FIELD_AND_PAGE_VISIBILITY_RULES_SDD §7.
_INGEST_MANIFEST: dict[RuleKind, str] = {
    RuleKind.VISIT_SCHEDULE: "VS rule (curated)",
    RuleKind.VALIDATION:     "Validation rule (curated)",
    RuleKind.EDIT_FORM:      "Edit form rule (curated)",
    RuleKind.DECISION:       "Decision rule (curated)",
    RuleKind.FORM_ELEMENT:   "Field rules (curated)",
}


# ── Porcelain: composite commands that orchestrate the surgical ones ─────────


def cmd_helpers(
    source: Path, root: Path, batch_size: int, skip_enrich: bool,
) -> int:
    """End-to-end helper refresh: sync → enrich-use-when → rebuild.

    Use after pulling new avni-models source or hand-editing helper files.
    Pass `--skip-enrich` to avoid the Haiku cost when you only want a fast
    sync + rebuild (e.g. when you'll annotate manually later).

    Each step exits non-zero on a hard failure; the chain stops at the first
    failing step so you can fix and retry without re-running successful work.
    """
    log.info("helpers: step 1/3 — sync from avni-models")
    rc = cmd_sync(source, root)
    if rc:
        log.error("helpers: sync failed; stopping")
        return rc

    if skip_enrich:
        log.info("helpers: step 2/3 — enrich-use-when SKIPPED (--skip-enrich)")
    else:
        log.info("helpers: step 2/3 — enrich-use-when via Haiku")
        rc = cmd_enrich_use_when(source, root, batch_size, dry_run=False)
        if rc:
            log.error("helpers: enrich-use-when failed; stopping")
            return rc

    log.info("helpers: step 3/3 — rebuild embedding cache")
    rc = cmd_rebuild(root)
    if rc:
        log.error("helpers: rebuild failed")
        return rc

    log.info("helpers: done")
    return 0


def cmd_examples(xlsx: Path, tab: str, root: Path, rule_kind: str) -> int:
    """End-to-end example refresh: ingest-examples → rebuild.

    Use after editing the curated rules xlsx tab. Cheap to re-run — the
    rebuild step is content-hash-based, so unchanged examples are not
    re-embedded.
    """
    log.info("examples: step 1/2 — ingest-examples from xlsx")
    rc = cmd_ingest_examples(xlsx, tab, root, rule_kind)
    if rc:
        log.error("examples: ingest failed; stopping")
        return rc

    log.info("examples: step 2/2 — rebuild embedding cache")
    rc = cmd_rebuild(root)
    if rc:
        log.error("examples: rebuild failed")
        return rc

    log.info("examples: done")
    return 0


def cmd_examples_all(xlsx: Path, root: Path) -> int:
    """Ingest curated tabs for every wired RuleKind and rebuild once.

    Iterates `_INGEST_MANIFEST`, calling `cmd_ingest_examples` for each
    rule kind / tab pair. Skips any kind whose tab is missing from the
    workbook (logged as a warning). Single `cmd_rebuild` at the end —
    avoids re-embedding the catalog four times.

    See FORM_LEVEL_RULES_SDD §7.
    """
    if not xlsx.exists():
        log.error(f"xlsx not found: {xlsx}")
        return 2

    total_kinds = len(_INGEST_MANIFEST)
    for i, (kind, tab) in enumerate(_INGEST_MANIFEST.items(), start=1):
        log.info(f"examples-all: step {i}/{total_kinds + 1} — {kind.value} from tab {tab!r}")
        rc = cmd_ingest_examples(xlsx, tab, root, kind.value)
        if rc:
            log.warning(
                f"examples-all: ingest failed for {kind.value!r} "
                f"(tab {tab!r}); continuing with remaining kinds"
            )

    log.info(f"examples-all: step {total_kinds + 1}/{total_kinds + 1} — rebuild embedding cache")
    rc = cmd_rebuild(root)
    if rc:
        log.error("examples-all: rebuild failed")
        return rc

    log.info("examples-all: done")
    return 0


# ── `sync` ────────────────────────────────────────────────────────────────────

# Maps avni-models class names to (catalog file name, JS variable used in signatures).
_ENTITY_CLASSES: list[tuple[str, str, str]] = [
    ("Individual",       "individual.json",        "individual"),
    ("ProgramEnrolment", "program_enrolment.json", "programEnrolment"),
    ("ProgramEncounter", "program_encounter.json", "programEncounter"),
    ("AbstractEncounter","abstract_encounter.json","encounter"),
    ("Observation",      "observation.json",       "observation"),
    ("Concept",          "concept.json",           "concept"),
]


def cmd_sync(source: Path, root: Path) -> int:
    """Walk avni-models JS files and update the per-entity helper catalogs."""
    if not source.exists():
        log.error(f"avni-models source not found: {source}")
        return 2

    entities_dir = root / "helpers" / "entities"
    entities_dir.mkdir(parents=True, exist_ok=True)

    total_methods = 0
    for class_name, file_name, var_name in _ENTITY_CLASSES:
        js_path = source / f"{class_name}.js"
        if not js_path.exists():
            log.warning(f"missing {js_path}; skipping {class_name}")
            continue
        try:
            methods = _extract_public_methods(js_path.read_text(encoding="utf-8"))
        except OSError as exc:
            log.warning(f"could not read {js_path}: {exc}")
            continue

        out_path = entities_dir / file_name
        existing = _load_existing_helpers(out_path)
        merged = _merge_helpers(class_name, var_name, methods, existing)
        out_path.write_text(json.dumps(merged, indent=2) + "\n", encoding="utf-8")
        log.info(f"{file_name}: {len(merged)} method(s) ({len(methods)} from source)")
        total_methods += len(merged)
    log.info(f"sync done — {total_methods} total method(s) across entities/")
    return 0


# Regexes that capture both the method name AND its parameter list.
# avni-models classes are mostly hand-rolled ES classes; the patterns below cover
# the declaration shapes that occur in `avni-models/src/<Class>.js`.
_METHOD_PATTERNS: list[re.Pattern[str]] = [
    # method(args) {
    re.compile(r"^\s*(?:static\s+)?(?:async\s+)?([a-zA-Z][\w$]*)\s*\(([^)]*)\)\s*\{"),
    # get prop()  — JS getter; no args, but it surfaces as a parameter-less call site
    re.compile(r"^\s*get\s+([a-zA-Z][\w$]*)\s*\(\s*\)"),
    # name = function(args)
    re.compile(r"^\s*([a-zA-Z][\w$]*)\s*=\s*function\s*\(([^)]*)\)"),
    # name = (args) =>
    re.compile(r"^\s*([a-zA-Z][\w$]*)\s*=\s*\(([^)]*)\)\s*=>"),
]
_SKIP_NAMES: frozenset[str] = frozenset({
    "constructor", "schema", "toString", "valueOf",
})
_RESERVED_WORDS: frozenset[str] = frozenset({
    "if", "for", "while", "switch", "return", "function",
    "class", "const", "let", "var", "throw", "import",
    "export", "new", "this", "default", "from", "case",
})


def _iter_method_declarations(js: str) -> Iterator[tuple[str, str, int]]:
    """Yield `(name, params, line_index)` for each method-like declaration.

    Walks the JS source line by line, applying `_METHOD_PATTERNS` and the
    skip-list (`_`-prefixed, `_SKIP_NAMES`, `_RESERVED_WORDS`). Comment-only
    lines (`//`, `*`, `/*`) are ignored. Consumers decide whether to keep
    every match (signature extraction) or whether to skip past captured
    method bodies (body extraction).
    """
    for idx, line in enumerate(js.splitlines()):
        stripped = line.lstrip()
        if stripped.startswith(("//", "*", "/*")):
            continue
        for pattern in _METHOD_PATTERNS:
            match = pattern.match(line)
            if not match:
                continue
            name = match.group(1)
            if name.startswith("_") or name in _SKIP_NAMES or name in _RESERVED_WORDS:
                break
            params = match.group(2).strip() if match.lastindex and match.lastindex >= 2 else ""
            yield name, params, idx
            break


def _extract_public_methods(js: str) -> list[tuple[str, str]]:
    """Extract `(name, parameter_list)` pairs from a JS class body.

    Best-effort regex — not a real JS parser. Parameter list is captured
    verbatim from the source (e.g. `'conceptNameOrUuid, currentEncounter'`)
    so the catalog can show realistic call sites.
    """
    seen: dict[str, str] = {}
    for name, params, _idx in _iter_method_declarations(js):
        seen.setdefault(name, params)
    return list(seen.items())


# Hand-curated annotations for the MVP set — the helpers SDD §6.1 names as
# required for day-one visit-schedule generation. Each entry overlays
# `use_when` + `example_snippet` on top of the regex-derived signature.
_MVP_ANNOTATIONS: dict[str, tuple[str, str]] = {
    # Individual
    "Individual.findLatestObservationFromPreviousEncounters": (
        "Read the latest observed value of a concept across encounters that "
        "occurred BEFORE `currentEncounter`. Look-back / trend rules.",
        "const last = individual.findLatestObservationFromPreviousEncounters('Risk', programEncounter);",
    ),
    "Individual.findLatestObservationInEntireEnrolment": (
        "Read the latest observed value of a concept across every encounter in "
        "every enrolment for this subject.",
        "const v = individual.findLatestObservationInEntireEnrolment('BP');",
    ),
    "Individual.hasEncounterOfType": (
        "True when at least one encounter of any of the given types exists.",
        "if (individual.hasEncounterOfType(['ANC'])) { /* ... */ }",
    ),
    "Individual.hasProgramEncounterOfType": (
        "True when at least one program-encounter of the given type exists.",
        "if (individual.hasProgramEncounterOfType('Delivery')) { /* ... */ }",
    ),
    "Individual.everScheduledEncountersOfType": (
        "All encounters of the given type ever scheduled for this subject "
        "(including completed and cancelled).",
        "const all = individual.everScheduledEncountersOfType('ANC');",
    ),
    "Individual.getEncounters": (
        "Subject's encounters. Pass `true` to include voided/cancelled.",
        "const encounters = individual.getEncounters(true);",
    ),
    # ProgramEnrolment
    "ProgramEnrolment.getEncounters": (
        "All encounters in this enrolment. Pass `true` to include voided/cancelled.",
        "const encs = programEnrolment.getEncounters(true);",
    ),
    "ProgramEnrolment.getEncountersOfType": (
        "All encounters of the given encounter-type name in this enrolment.",
        "const followups = programEnrolment.getEncountersOfType('Followup');",
    ),
    "ProgramEnrolment.hasEncounter": (
        "True when an encounter of the given name exists in this enrolment.",
        "if (programEnrolment.hasEncounter('SCD Followup Jan 2024')) { return scheduleBuilder.getAll(); }",
    ),
    "ProgramEnrolment.hasCompletedEncounterOfType": (
        "True when an encounter of the given type has been completed (has "
        "`encounterDateTime`).",
        "if (!programEnrolment.hasCompletedEncounterOfType('Baseline')) { /* schedule it */ }",
    ),
    "ProgramEnrolment.scheduledEncountersOfType": (
        "Encounters of the given type that are scheduled but not yet completed.",
        "const pending = programEnrolment.scheduledEncountersOfType('Followup');",
    ),
    "ProgramEnrolment.getObservationReadableValueInEntireEnrolment": (
        "Read the human-readable observed value of a concept across the whole "
        "enrolment.",
        "const result = programEnrolment.getObservationReadableValueInEntireEnrolment('Electrophoresis result');",
    ),
    # ProgramEncounter (delegates mostly to ProgramEnrolment)
    "ProgramEncounter.getObservationReadableValue": (
        "Read the human-readable value of a concept observed on THIS program "
        "encounter. First-arg string is the concept name.",
        "const risk = programEncounter.getObservationReadableValue('Risk');",
    ),
    # AbstractEncounter
    "AbstractEncounter.getObservationReadableValue": (
        "Read the human-readable value of a concept observed on this encounter.",
        "const bt = encounter.getObservationReadableValue('BT date');",
    ),
    "AbstractEncounter.findCancelEncounterObservationReadableValue": (
        "Read the human-readable value of a concept recorded on the "
        "cancellation form of this encounter (only on cancellation rule kinds).",
        "const reason = programEncounter.findCancelEncounterObservationReadableValue('Visit cancel reason');",
    ),
    "AbstractEncounter.getRealEventDate": (
        "Return `earliestVisitDateTime || encounterDateTime` — handles upload "
        "cases where earliest is null. Prefer over the inline `||` idiom.",
        "const visitDate = encounter.getRealEventDate();",
    ),
}


def _annotation_for(name: str) -> tuple[str, str]:
    """Return `(use_when, example_snippet)` for a method, empty when none."""
    return _MVP_ANNOTATIONS.get(name, ("", ""))


def _load_existing_helpers(path: Path) -> dict[str, dict[str, Any]]:
    """Read prior catalog (if any) keyed by `name` so merges preserve hand edits."""
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except OSError as exc:
        log.warning(f"could not read {path}: {exc}")
        return {}
    except json.JSONDecodeError as exc:
        log.warning(f"{path} is not valid JSON ({exc}); ignoring prior catalog")
        return {}
    if isinstance(data, list):
        return {entry["name"]: entry for entry in data if entry.get("name")}
    return {}


def _merge_helpers(
    class_name: str,
    var_name: str,
    methods: list[tuple[str, str]],
    existing: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Combine source-derived `(name, params)` pairs with the existing catalog.

    Precedence on each field:
      1. Prior hand-written value wins.
      2. Otherwise, MVP annotation overlay (see _MVP_ANNOTATIONS).
      3. Otherwise, regex-derived value (or empty).
    """
    merged: list[dict[str, Any]] = []
    seen: set[str] = set()
    for method, params in methods:
        name = f"{class_name}.{method}"
        seen.add(name)
        prior = existing.get(name, {})
        annot_use_when, annot_snippet = _annotation_for(name)
        derived_signature = f"{var_name}.{method}({params})" if params else f"{var_name}.{method}()"
        entry: dict[str, Any] = {
            "name": name,
            "signature": prior.get("signature") or derived_signature,
            "use_when": prior.get("use_when") or annot_use_when,
            "example_snippet": prior.get("example_snippet") or annot_snippet,
        }
        # `applies_to` is optional: omitted = applies to every RuleKind. Only
        # carry it forward when the prior catalog narrowed scope intentionally
        # (e.g. `imports/visit_schedule.json`).
        if prior.get("applies_to"):
            entry["applies_to"] = prior["applies_to"]
        merged.append(entry)
    # Keep any hand-written rows whose method is no longer detectable in source —
    # they may have been added intentionally despite a non-trivial declaration.
    for name, entry in existing.items():
        if name not in seen:
            merged.append(entry)
    merged.sort(key=lambda e: e["name"])
    return merged


# ── `ingest-examples` ─────────────────────────────────────────────────────────


@dataclass
class _IngestStats:
    written: int = 0
    skipped: int = 0


def cmd_ingest_examples(xlsx: Path, tab: str, root: Path, rule_kind: str) -> int:
    """Read the curated rules tab and write one Markdown file per row.

    Columns are resolved by header name (case-insensitive). The curated
    tabs vary in shape — form-level tabs carry
    ``ORG name | Form name | Rule | Prompt`` while the Field-rules tab
    adds ``field_name`` and ``kind``. Both are handled by the same loop.
    """
    if not xlsx.exists():
        log.error(f"xlsx not found: {xlsx}")
        return 2
    try:
        workbook = openpyxl.load_workbook(xlsx, data_only=True)
    except Exception as exc:  # noqa: BLE001
        log.error(f"could not open {xlsx}: {exc}")
        return 2
    if tab not in workbook.sheetnames:
        log.error(f"tab {tab!r} not in workbook (have: {workbook.sheetnames})")
        return 2

    out_dir = root / "examples" / rule_kind
    out_dir.mkdir(parents=True, exist_ok=True)

    sheet = workbook[tab]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        log.warning(f"tab {tab!r} is empty")
        return 0

    column_indices = _resolve_column_indices(header_row)
    if column_indices.get("rule") is None or column_indices.get("prompt") is None:
        log.error(
            f"tab {tab!r} missing required 'rule' or 'prompt' header "
            f"(have: {list(header_row)})"
        )
        return 2

    stats = _IngestStats()
    for row in rows_iter:
        if not row or not any(cell for cell in row if cell):
            continue
        org        = _cell_text(row, column_indices.get("org_name"))
        form       = _cell_text(row, column_indices.get("form_name"))
        field_name = _cell_text(row, column_indices.get("field_name"))
        rule_js    = _cell_text(row, column_indices["rule"])
        prompt     = _cell_text(row, column_indices["prompt"])
        kind       = _cell_text(row, column_indices.get("kind"))

        if not rule_js or not prompt:
            stats.skipped += 1
            log.warning(
                f"row skipped (org={org!r} form={form!r} "
                f"field={field_name!r}): empty rule or prompt"
            )
            continue

        entity_param = _derive_entity_param(rule_js) or "individual"
        encounter_types = _derive_encounter_types(rule_js)
        concepts = _derive_concepts(rule_js)

        slug_basis = f"{org}_{form}_{field_name}" if field_name else f"{org}_{form}"
        slug = _slug(slug_basis)
        path = out_dir / f"{slug}.md"
        path.write_text(
            _render_example(
                rule_kind, prompt, entity_param, encounter_types,
                concepts, org, rule_js,
                field_name=field_name or None,
                kind=kind or None,
            ),
            encoding="utf-8",
        )
        stats.written += 1

    log.info(f"ingest-examples done — wrote {stats.written}, skipped {stats.skipped}")
    return 0


def _resolve_column_indices(header_row: tuple) -> dict[str, int]:
    """Build a `{canonical_key → column_index}` map from the header row.

    Canonical keys are: ``org_name``, ``form_name``, ``field_name``,
    ``rule``, ``prompt``, ``kind``. Headers are matched case-insensitively
    against the known synonym table below; columns whose header maps to no
    known key are ignored. Missing columns return None for that key.
    """
    aliases: dict[str, tuple[str, ...]] = {
        "org_name":   ("org_name", "org name"),
        "form_name":  ("form_name", "form name"),
        "field_name": ("field_name", "field name"),
        "rule":       ("rule",),
        "prompt":     ("prompt",),
        "kind":       ("kind",),
    }
    out: dict[str, int] = {}
    for col, raw in enumerate(header_row):
        if raw is None:
            continue
        header_text = str(raw).strip().lower()
        if not header_text:
            continue
        for canonical, synonyms in aliases.items():
            if canonical in out:
                continue
            if header_text in synonyms:
                out[canonical] = col
                break
    return out


def _cell_text(row: tuple, idx: int | None) -> str:
    """Read a cell as stripped text. None / missing column → empty string."""
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    return str(value).strip()


_PARAM_ENTITY_RE = re.compile(r"const\s+(\w+)\s*=\s*params\.entity")
_VS_BUILDER_RE = re.compile(r"VisitScheduleBuilder\s*\(\s*\{\s*(\w+)")
_ENCOUNTER_TYPE_PROPERTY_RE = re.compile(
    r"""encounterType\s*:\s*['"]([^'"]+)['"]""", re.DOTALL,
)
_CONCEPT_ACCESSOR_RE = concept_accessor_call_regex()


def _derive_entity_param(js: str) -> str:
    """Find the JS variable bound to `params.entity` (or the builder arg)."""
    match = _PARAM_ENTITY_RE.search(js)
    if match:
        return match.group(1)
    match = _VS_BUILDER_RE.search(js)
    if match:
        return match.group(1)
    return ""


def _derive_encounter_types(js: str) -> list[str]:
    """Collect encounter-type names from `encounterType: '...'` literals."""
    return sorted(set(_ENCOUNTER_TYPE_PROPERTY_RE.findall(js)))


def _derive_concepts(js: str) -> list[str]:
    """Collect concept names from observation-accessor first-arg literals."""
    return sorted(set(_CONCEPT_ACCESSOR_RE.findall(js)))


_SLUG_RE = re.compile(r"[^a-z0-9]+")


def _slug(text: str) -> str:
    s = _SLUG_RE.sub("_", text.lower()).strip("_")
    return s or "example"


def _render_example(
    rule_kind: str,
    intent: str,
    entity_param: str,
    encounter_types: list[str],
    concepts: list[str],
    source_org: str,
    js: str,
    *,
    field_name: str | None = None,
    kind: str | None = None,
) -> str:
    """Render a single example Markdown file matching VISIT_SCHEDULE_RULE_SDD §6.2.

    Field-rule examples (FIELD_AND_PAGE_VISIBILITY_RULES_SDD §7) also carry
    the originating ``field_name`` and the behaviour ``kind`` tag
    (``visibility | value | validation | answerFilter``). Both are
    informational — the embedder still vectorises only ``intent`` — so
    callers can omit them for the form-level corpora.
    """
    payload: dict[str, Any] = {
        "rule_kind": rule_kind,
        "intent": intent,
        "entity_param": entity_param,
        "encounter_types": encounter_types,
        "concepts": concepts,
        "source_org": source_org,
    }
    if field_name:
        payload["field_name"] = field_name
    if kind:
        payload["kind"] = kind

    frontmatter = yaml.safe_dump(
        payload,
        sort_keys=False,
        allow_unicode=True,
        default_flow_style=False,
    )
    return f"---\n{frontmatter}---\n```js\n{js}\n```\n"


# ── `rebuild` ─────────────────────────────────────────────────────────────────


def cmd_rebuild(root: Path) -> int:
    """Force a full re-embed of the catalog into `.embeddings.json`."""
    cache_path = root / ".embeddings.json"
    if cache_path.exists():
        cache_path.unlink()
    kb = KnowledgeBase(root=root)
    if not kb.helpers and not kb.examples:
        log.error("catalog is empty — populate helpers/ and examples/ first")
        return 2
    try:
        # `_ensure_catalog_vectors` walks every entry and embeds those missing
        # from the cache — which, post-unlink, is all of them.
        kb._ensure_catalog_vectors()  # noqa: SLF001
    except RuntimeError as exc:
        log.error(str(exc))
        return 2
    log.info(f"rebuild done — wrote {cache_path}")
    return 0


# ── `enrich-use-when` ─────────────────────────────────────────────────────────


class _MethodAnnotation(BaseModel):
    """One Haiku-authored annotation for a single method."""

    method: str
    use_when: str


class _ClassAnnotations(BaseModel):
    """Structured output schema for the per-batch Haiku call."""

    annotations: list[_MethodAnnotation] = Field(default_factory=list)


_ENRICH_SYSTEM_PROMPT = """\
You are documenting JavaScript methods on an avni-models class for an internal
rule-authoring catalog. For every method you receive, write ONE sentence
(maximum 20 words) describing when an Avni rule (visit schedule, validation,
eligibility, summary) would call it.

Rules for the annotations:
  - Lead with a verb in present tense (e.g. "Reads...", "Returns true when...",
    "Schedules..."). No "this method".
  - Be concrete. Reference what the method returns or mutates.
  - If a method is purely internal (CRUD mutation, serialization, framework
    plumbing, validation of registration fields), say so directly — e.g.
    "Appends an encounter — internal mutation, rarely called from rules."
  - Never invent parameters or behaviour the body doesn't show.
  - Never reference the class name or `params.entity` — readers know the context.

Return exactly one annotation per method, preserving the order received.
"""


def cmd_enrich_use_when(
    source: Path, root: Path, batch_size: int, dry_run: bool,
) -> int:
    """Fill empty `use_when` fields by asking Haiku to read each method body."""
    if not os.environ.get("ANTHROPIC_API_KEY"):
        log.error("ANTHROPIC_API_KEY is not set")
        return 2
    if not source.exists():
        log.error(f"avni-models source not found: {source}")
        return 2

    entities_dir = root / "helpers" / "entities"
    if not entities_dir.exists():
        log.error(f"entities catalog not found at {entities_dir} — run sync first")
        return 2

    try:
        from langchain_anthropic import ChatAnthropic
        from langchain_core.messages import HumanMessage, SystemMessage
    except ImportError as exc:
        log.error(f"langchain-anthropic not installed: {exc}")
        return 2

    chat = ChatAnthropic(model="claude-haiku-4-5", max_tokens=4096)
    structured = chat.with_structured_output(_ClassAnnotations)

    grand_total = 0
    for class_name, file_name, _var in _ENTITY_CLASSES:
        js_path = source / f"{class_name}.js"
        catalog_path = entities_dir / file_name
        if not js_path.exists() or not catalog_path.exists():
            log.warning(f"skipping {class_name}: missing source or catalog")
            continue

        try:
            entries = json.loads(catalog_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            log.warning(f"could not read {catalog_path}: {exc}")
            continue
        if not isinstance(entries, list):
            log.warning(f"{catalog_path} is not a JSON array — skipping")
            continue

        bodies = _extract_method_bodies(js_path.read_text(encoding="utf-8"))
        targets: list[tuple[int, str, str]] = []  # (entry_index, method, body)
        for idx, entry in enumerate(entries):
            if entry.get("use_when"):
                continue
            name = entry.get("name", "")
            method = name.split(".", 1)[1] if "." in name else name
            body = bodies.get(method)
            if not body:
                continue
            targets.append((idx, method, body))

        if not targets:
            log.info(f"{file_name}: nothing to enrich")
            continue

        log.info(f"{file_name}: annotating {len(targets)} method(s)")
        for batch_start in range(0, len(targets), batch_size):
            batch = targets[batch_start:batch_start + batch_size]
            user_text = _render_enrich_user_prompt(class_name, batch)
            messages = [
                SystemMessage(content=[{
                    "type": "text",
                    "text": _ENRICH_SYSTEM_PROMPT,
                    "cache_control": {"type": "ephemeral"},
                }]),
                HumanMessage(content=user_text),
            ]
            try:
                result: _ClassAnnotations = structured.invoke(messages)  # type: ignore[assignment]
            except Exception as exc:  # noqa: BLE001
                log.warning(f"Haiku call failed for {class_name} batch "
                            f"{batch_start}: {exc}")
                continue

            by_method = {a.method: a.use_when.strip() for a in result.annotations}
            for entry_idx, method, _body in batch:
                annotation = by_method.get(method)
                if not annotation:
                    continue
                entries[entry_idx]["use_when"] = annotation
                grand_total += 1

        if dry_run:
            log.info(f"{file_name}: dry-run — not writing")
        else:
            catalog_path.write_text(
                json.dumps(entries, indent=2) + "\n", encoding="utf-8",
            )

    log.info(f"enrich-use-when done — populated {grand_total} annotation(s)")
    return 0


def _render_enrich_user_prompt(
    class_name: str, batch: list[tuple[int, str, str]],
) -> str:
    """Build the user message for one Haiku batch."""
    parts = [f"Class: {class_name}", ""]
    for _idx, method, body in batch:
        parts.append(f"### {method}")
        parts.append("```js")
        parts.append(body)
        parts.append("```")
        parts.append("")
    return "\n".join(parts)


def _extract_method_bodies(js: str) -> dict[str, str]:
    """Return `{method_name: body_text}` for every top-level method in a class.

    Best-effort brace-balancing extractor — not a real JS parser, but the
    avni-models classes are flat enough that this works in practice. Matches
    that fall inside an already-captured body are skipped so nested
    function-like syntax isn't recorded as a separate method.
    """
    lines = js.splitlines()
    out: dict[str, str] = {}
    skip_until = -1
    for name, _params, idx in _iter_method_declarations(js):
        if idx <= skip_until:
            continue
        if name in out:
            # Keep the first definition; later ones could be call-sites the
            # permissive regex matched.
            continue
        body_lines, end = _collect_balanced_body(lines, idx)
        if body_lines:
            out[name] = "\n".join(body_lines)
            skip_until = end
    return out


def _collect_balanced_body(
    lines: list[str], start: int,
) -> tuple[list[str], int]:
    """Capture lines from `start` until the method's outer `{...}` closes.

    Returns the captured lines plus the index of the closing line. If no brace
    balance is reached (truncated source / arrow-no-body), returns ([], start).
    """
    depth = 0
    seen_open = False
    captured: list[str] = []
    for j in range(start, len(lines)):
        line = lines[j]
        captured.append(line)
        # Track depth across `{` and `}` characters, ignoring those inside
        # string / template literals via a light state machine.
        depth += _net_brace_delta(line)
        if not seen_open and "{" in line:
            seen_open = True
        if seen_open and depth == 0:
            return captured, j
    return [], start


def _net_brace_delta(line: str) -> int:
    """Net `{` - `}` count for a single line, ignoring quoted braces."""
    delta = 0
    i = 0
    n = len(line)
    in_str: str | None = None  # one of "'\"`" when inside a string literal
    while i < n:
        ch = line[i]
        if in_str:
            if ch == "\\":
                i += 2
                continue
            if ch == in_str:
                in_str = None
        else:
            if ch in ("'", '"', "`"):
                in_str = ch
            elif ch == "/" and i + 1 < n and line[i + 1] == "/":
                break  # rest is a line comment
            elif ch == "{":
                delta += 1
            elif ch == "}":
                delta -= 1
        i += 1
    return delta


if __name__ == "__main__":
    sys.exit(main())
