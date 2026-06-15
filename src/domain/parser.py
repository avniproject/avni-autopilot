"""
scoping_parser.py — Parse any combination of SRS/scoping/modelling files into
a validated AVNI EntitySpec.

Accepts: .xlsx (multi-sheet) — any number of files.
Each sheet is auto-classified by its content (column headers, data patterns)
into one of: location_hierarchy, subject_types, programs, encounters,
program_encounters, w3h_mapping, form_fields, or ignored.

The parser is content-driven, not name-driven — it works regardless of sheet
names or column header variations.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd

from domain.rules.parser import (
    apply_intents_to_forms,
    detect_rule_columns,
    extract_rule_intents,
    find_form_name_column,
)
from models import (
    AddressLevelSpec,
    EncounterTypeSpec,
    EntitySpec,
    FieldSpec,
    FormSpec,
    GroupSpec,
    ProgramSpec,
    SectionSpec,
    SkipLogicSpec,
    SubjectTypeSpec,
)

pd.set_option("future.no_silent_downcasting", True)

logger = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────

_SUBJECT_TYPE_MAP = {
    "individual": "Person",
    "person": "Person",
    "group": "Group",
    "household": "Household",
}

_DATA_TYPE_MAP: dict[str, str] = {
    "text": "Text",
    "string": "Text",
    "varchar": "Text",
    "numeric": "Numeric",
    "number": "Numeric",
    "integer": "Numeric",
    "int": "Numeric",
    "decimal": "Numeric",
    "float": "Numeric",
    "date": "Date",
    "datetime": "DateTime",
    "pre added options": "Coded",
    "coded": "Coded",
    "single select": "Coded",
    "single-select": "Coded",
    "multi select": "Coded",
    "multi-select": "Coded",
    "dropdown": "Coded",
    "radio": "Coded",
    "checkbox": "Coded",
    "boolean": "Coded",
    "bool": "Coded",
    "yes/no": "Coded",
    "subject": "Subject",
    "duration": "Duration",
    "notes": "Notes",
    "image": "ImageV2",
    "imagev2": "ImageV2",
    "photo": "ImageV2",
    "phone number": "PhoneNumber",
    "phonenumber": "PhoneNumber",
    "phone": "PhoneNumber",
    "mobile": "PhoneNumber",
    "location": "Location",
    "file": "File",
    "audio": "Audio",
    "video": "Video",
    "questiongroup": "QuestionGroup",
    "question group": "QuestionGroup",
    "repeatable": "QuestionGroup",
}

# Header patterns for auto-classification (lowercase, checked via substring)
_LOCATION_HEADERS = {"location type", "location", "address level", "hierarchy"}
_SUBJECT_HEADERS = {"subject type", "subject type name", "entity type", "type"}
_PROGRAM_HEADERS = {"program name", "program", "target subject"}
_ENCOUNTER_HEADERS = {"encounter name", "encounter type", "scheduled", "unscheduled"}
_FORM_HEADERS = {"field name", "data type", "mandatory", "page name"}
_W3H_HEADERS = {"what", "when", "who", "how"}


# ── Helpers ──────────────────────────────────────────────────────────────────


def _clean(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


# Regex to strip leading numbering prefixes from field/concept names:
# "1. Name" → "Name", "2) Age" → "Age", "A. Gender" → "Gender",
# "1.2 Weight" → "Weight", "1.2.3 Height" → "Height"
_NUMBER_PREFIX_RE = re.compile(
    r"^\s*"
    r"(?:"
    r"[0-9]+(?:\.[0-9]+)*"  # "1", "1.2", "1.2.3"
    r"|[A-Za-z]"  # single letter like "A", "B"
    r")"
    r"[\.\)\-:\s]\s*"  # followed by . ) - : or space
)


def _clean_field_name(val: Any) -> str:
    """Clean a field/concept name: strip whitespace and numbering prefixes."""
    name = _clean(val)
    if not name:
        return name
    # Strip leading number prefixes like "1. ", "2) ", "A. ", "1.2 "
    name = _NUMBER_PREFIX_RE.sub("", name)
    # Also strip trailing numbering artefacts
    return name.strip()


def _parse_yes_no(val: Any) -> bool:
    return _clean(val).lower() in ("yes", "y", "true", "1")


def _parse_options(val: Any) -> list[str]:
    raw = _clean(val)
    if not raw:
        return []
    if "\n" in raw:
        parts = raw.split("\n")
    elif ";" in raw:
        parts = raw.split(";")
    else:
        parts = raw.split(",")
    return [
        _NUMBER_PREFIX_RE.sub("", p.strip().rstrip(",")).strip()
        for p in parts
        if p.strip() and p.strip().lower() != "nan"
    ]




def _parse_min_max(val: Any) -> tuple[float | None, float | None]:
    raw = _clean(val)
    if not raw:
        return None, None
    m = re.match(r"(-?[\d.]+)\s*[-–to]+\s*(-?[\d.]+)", raw, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1)), float(m.group(2))
        except ValueError:
            pass
    return None, None


def _fuzzy_match(
    query: str, candidates: set[str] | list[str], threshold: float = 0.5
) -> str | None:
    """Match a query string against candidates using exact, substring, then word overlap.

    Returns the best matching candidate name, or None if no match above threshold.
    Uses word-level overlap (not character-level) to avoid false matches
    between strings that share common letters but different meanings.
    Handles typos via character scoring only for short single-word names.
    """
    if not query:
        return None
    q = query.strip().lower()
    if not q:
        return None

    # 1. Exact match
    for c in candidates:
        if c.lower() == q:
            return c

    # 2. Substring containment (longer overlap = better)
    best_sub = None
    best_sub_len = 0
    for c in candidates:
        cl = c.lower()
        if cl in q or q in cl:
            overlap = min(len(cl), len(q))
            if overlap > best_sub_len:
                best_sub = c
                best_sub_len = overlap
    if best_sub:
        return best_sub

    q_words = set(re.findall(r"[a-z0-9]+", q))
    if not q_words:
        return None

    # 3. Word overlap scoring (semantic matching)
    best_word_match = None
    best_word_score = 0.0
    for c in candidates:
        cl = c.lower()
        c_words = set(re.findall(r"[a-z0-9]+", cl))
        if not c_words:
            continue
        common_words = q_words & c_words
        # Jaccard similarity on words
        score = len(common_words) / len(q_words | c_words)
        if score > threshold and score > best_word_score:
            best_word_match = c
            best_word_score = score

    if best_word_match:
        return best_word_match

    # 4. Character overlap (only for short names — handles typos like Paritcipant)
    # Skip for multi-word names to avoid false matches
    if len(q_words) <= 2:
        best_fuzzy = None
        best_char_score = 0.0
        for c in candidates:
            cl = c.lower()
            c_words_count = len(re.findall(r"[a-z0-9]+", cl))
            if c_words_count > 2:
                continue  # Don't char-match multi-word candidates
            if not cl:
                continue
            common = sum(1 for ch in q if ch in cl)
            score = common / max(len(q), len(cl))
            if score > 0.7 and score > best_char_score:
                best_fuzzy = c
                best_char_score = score
        if best_fuzzy:
            return best_fuzzy

    return None


def _resolve_subject_type(raw: str, known: set[str]) -> str:
    if not raw:
        return ""
    # Try raw name first, then cleaned version
    match = _fuzzy_match(raw, known)
    if match:
        return match
    cleaned = _clean_subject_name(raw)
    if cleaned.lower() != raw.strip().lower():
        match = _fuzzy_match(cleaned, known)
        if match:
            return match
    return _clean_subject_name(raw)


def _map_data_type(raw: str) -> str:
    lower = raw.strip().lower()
    return _DATA_TYPE_MAP.get(lower, "Text")


def _find_col(df: pd.DataFrame, *candidates: str) -> str | None:
    """Find a column in df by fuzzy matching against candidate names."""
    cols_lower = {str(c).strip().lower(): str(c) for c in df.columns}
    for cand in candidates:
        cand_lower = cand.lower()
        # Exact
        if cand_lower in cols_lower:
            return cols_lower[cand_lower]
        # Substring
        for col_lower, col_orig in cols_lower.items():
            if cand_lower in col_lower or col_lower in cand_lower:
                return col_orig
    return None


# ── Sheet Classification ─────────────────────────────────────────────────────


def _classify_sheet(df: pd.DataFrame, sheet_name: str = "") -> str:
    """
    Classify a DataFrame by its column headers into one of:
    location, subject_type, program, encounter, program_encounter,
    w3h, form, or unknown.

    Uses the FIRST column as the strongest signal — it's usually the primary
    entity name (e.g. "Encounter Name", "Subject Type Name", "Program Name").
    Falls back to sheet_name pattern matching when content headers are ambiguous.
    """
    if df.empty or df.shape[1] < 2:
        return "unknown"

    cols_lower = [str(c).strip().lower() for c in df.columns]
    cols_set = set(cols_lower)
    first_col = cols_lower[0] if cols_lower else ""

    # Unified modelling sheet — "Type" as first column with entity type values
    if first_col == "type" and "name" in cols_set:
        # Check if data rows contain entity type values
        type_vals = set()
        for _, row in df.head(10).iterrows():
            v = str(row.iloc[0]).strip().lower()
            if v in ("subject type", "program", "encounter type", "address level"):
                type_vals.add(v)
        if len(type_vals) >= 2:
            return "unified_modelling"

    # W3H — "what" as first column + "when"/"who"/"how"
    if first_col in ("what", "activity") and any(
        c in cols_set for c in ("when", "who", "how")
    ):
        return "w3h"

    # Form fields — "field name" or "page name" + "data type"
    has_field_name = any("field name" in c for c in cols_lower)
    has_data_type = any("data type" in c or "datatype" in c for c in cols_lower)
    if has_field_name and has_data_type:
        return "form"

    # Form-level rules tab — name column ("form name" / "form" / "name") plus
    # at least one rule-column alias. Detected BEFORE "form" so a sheet that
    # only carries rules doesn't get mistaken for a field-list sheet.
    if find_form_name_column(df) is not None and detect_rule_columns(df):
        return "rules"

    # Encounters — first col contains "encounter name" or "encounter type"
    if (
        "encounter name" in first_col
        or "encounter type" in first_col
        or first_col in ("encounter", "encounter type")
    ):
        # Program encounters have a "program" column too
        if any("program" in c for c in cols_lower[1:]):
            return "program_encounter"
        return "encounter"

    # Subject types — first col is "subject type name" or similar
    if "subject type name" in first_col or first_col == "subject type":
        return "subject_type"

    # Programs — first col is "program name" or standalone "program" with structural cols
    if "program name" in first_col or first_col == "program name":
        return "program"
    if first_col == "program" and any(
        "target" in c
        or "subject" in c
        or "enrolment" in c
        or "colour" in c
        or "color" in c
        for c in cols_lower
    ):
        return "program"

    # Location hierarchy — first col is "location type" or similar
    if "location" in first_col and (
        "type" in first_col or "hierarchy" in first_col or "level" in first_col
    ):
        return "location"
    if first_col == "location type":
        return "location"

    # ── Sheet-name fallback (when column headers are ambiguous) ─────────────
    if sheet_name:
        sn = sheet_name.strip().lower()
        if "program encounter" in sn:
            return "program_encounter"
        if sn in ("program", "programs"):
            return "program"
        if "encounter" in sn and "program" not in sn:
            return "encounter"
        if "subject type" in sn or sn in ("subject types", "subjects"):
            return "subject_type"
        if "location hierarchy" in sn or sn in (
            "location",
            "locations",
            "address levels",
        ):
            return "location"

    return "unknown"


# ── Entity Parsers ───────────────────────────────────────────────────────────


def parse_location_hierarchy(df: pd.DataFrame) -> list[AddressLevelSpec]:
    levels: list[AddressLevelSpec] = []
    rows = [(_clean(r[0]),) for r in df.itertuples(index=False) if _clean(r[0])]
    total = len(rows)
    prev_name: str | None = None
    for i, (loc_type,) in enumerate(rows):
        level = total - i
        levels.append(AddressLevelSpec(name=loc_type, level=level, parent=prev_name))
        prev_name = loc_type
    return levels


_SUBJECT_NAME_STRIP_SUFFIXES = [
    " registration",
    " enrolment",
    " enrollment",
    " form",
    " profile",
    " details",
    " entry",
]

# Header row values that should be filtered out
_HEADER_ROW_VALUES = {
    "subject type",
    "subject type name",
    "name",
    "entity name",
    "encounter name",
    "encounter",
    "program name",
    "program",
    "type",
    "entity type",
    "scheduled",
    "unscheduled",
}


def _clean_subject_name(raw: str) -> str:
    """Clean a subject type name: strip form-like suffixes.
    'Beneficiary Registration' → 'Beneficiary'
    'School Registration' → 'School'
    """
    name = raw.strip()
    lower = name.lower()
    for suffix in _SUBJECT_NAME_STRIP_SUFFIXES:
        if lower.endswith(suffix):
            name = name[: len(name) - len(suffix)].strip()
            break
    return name


def _is_header_row(name: str) -> bool:
    """Check if a parsed name is actually a header row value."""
    return name.strip().lower() in _HEADER_ROW_VALUES


def parse_subject_types(df: pd.DataFrame) -> list[SubjectTypeSpec]:
    subject_types: list[SubjectTypeSpec] = []
    seen: set[str] = set()

    name_col = _find_col(df, "Subject Type Name", "Subject Type", "Name", "Entity Name")
    type_col = _find_col(df, "Type", "Subject Type", "Entity Type")

    if not name_col:
        return []

    for _, row in df.iterrows():
        raw_name = _clean(row.get(name_col, ""))
        if not raw_name or _is_header_row(raw_name):
            continue

        name = _clean_subject_name(raw_name)
        if not name or name.lower() in seen:
            continue
        seen.add(name.lower())

        raw_type = _clean(row.get(type_col, "Person")).lower() if type_col else "person"
        avni_type = _SUBJECT_TYPE_MAP.get(raw_type, "Person")

        subject_types.append(
            SubjectTypeSpec(
                name=name,
                type=avni_type,
                allowProfilePicture=False,
                uniqueName=False,
                lastNameOptional=True,
            )
        )
    return subject_types


def parse_programs(df: pd.DataFrame, subject_type_names: set[str]) -> list[ProgramSpec]:
    programs: list[ProgramSpec] = []
    seen: set[str] = set()

    name_col = _find_col(df, "Program Name", "Program", "Name")
    target_col = _find_col(df, "Target Subject Type", "Target Subject", "Subject Type")

    if not name_col:
        return []

    for _, row in df.iterrows():
        name = _clean(row.get(name_col, ""))
        if not name or name.lower() in seen:
            continue
        seen.add(name.lower())

        target = _clean(row.get(target_col, "")) if target_col else ""
        if target and subject_type_names:
            # Try fuzzy match against known subject types
            matched = _fuzzy_match(target, subject_type_names)
            if matched:
                target = matched
            elif len(subject_type_names) == 1:
                # Single subject type — assume it's the target
                target = next(iter(subject_type_names))
        elif not target and subject_type_names:
            target = next(iter(subject_type_names))

        programs.append(
            ProgramSpec(name=name, target_subject_type=target, colour="#4A148C")
        )
    return programs


def parse_encounters(
    df: pd.DataFrame,
    subject_type_names: set[str],
    program_names: set[str],
    is_program_encounter: bool = False,
) -> list[EncounterTypeSpec]:
    encounter_types: list[EncounterTypeSpec] = []
    seen: set[str] = set()

    name_col = _find_col(df, "Encounter Name", "Encounter", "Name")
    if not name_col:
        return []

    sched_col = _find_col(
        df,
        "Encounter Type (Scheduled/Unscheduled)",
        "Encounter Type",
        "Scheduled",
        "Type",
    )

    if is_program_encounter:
        prog_col = _find_col(df, "Program name", "Program Name", "Program")
        st_col = None
    else:
        prog_col = None
        st_col = _find_col(df, "Subject Type", "Subject", "Entity")

    for _, row in df.iterrows():
        name = _clean(row.get(name_col, ""))
        if not name or name.lower() in seen:
            continue
        # Skip header rows that got parsed as data
        if _is_header_row(name):
            continue
        seen.add(name.lower())

        program_name = ""
        subject_type = ""

        if is_program_encounter and prog_col:
            raw_prog = _clean(row.get(prog_col, ""))
            if not _is_header_row(raw_prog):
                program_name = raw_prog
        elif st_col:
            raw_subject = _clean(row.get(st_col, ""))
            if not _is_header_row(raw_subject):
                subject_type = _resolve_subject_type(raw_subject, subject_type_names)

        enc_type_raw = ""
        if sched_col:
            enc_type_raw = _clean(row.get(sched_col, "")).lower()
        is_scheduled = "unscheduled" not in enc_type_raw

        encounter_types.append(
            EncounterTypeSpec(
                name=name,
                program_name=program_name,
                subject_type=subject_type,
                is_program_encounter=is_program_encounter or bool(program_name),
                is_scheduled=is_scheduled,
            )
        )
    return encounter_types


# ── W3H Parser ───────────────────────────────────────────────────────────────


def parse_w3h(
    df: pd.DataFrame,
    subject_type_names: set[str],
    encounter_type_names: set[str],
    program_encounter_map: dict[str, str],
) -> dict[str, dict[str, str | None]]:
    """Parse W3H sheet to build form → entity mapping."""
    what_col = _find_col(df, "What", "Activity", "Form")
    if not what_col:
        what_col = str(df.columns[0]) if len(df.columns) > 0 else None
    if not what_col:
        return {}

    st_names_lower = {n.lower(): n for n in subject_type_names}
    enc_names_lower = {n.lower(): n for n in encounter_type_names}

    _REG_KEYWORDS = (
        "registration",
        "register",
        "details",
        "profile",
        "enrolment",
        "enrollment",
    )

    mapping: dict[str, dict[str, str | None]] = {}
    for _, row in df.iterrows():
        activity = _clean(row.get(what_col, ""))
        if not activity:
            continue

        activity_lower = activity.lower()
        entry: dict[str, str | None] = {
            "formType": None,
            "subjectType": None,
            "encounterType": None,
            "program": None,
        }

        is_registration = any(kw in activity_lower for kw in _REG_KEYWORDS)

        if is_registration:
            entry["formType"] = "IndividualProfile"
            for st_lower, st_name in st_names_lower.items():
                if st_lower in activity_lower or activity_lower in st_lower:
                    entry["subjectType"] = st_name
                    break
            if not entry["subjectType"] and len(subject_type_names) == 1:
                entry["subjectType"] = next(iter(subject_type_names))
        else:
            matched_enc = None
            for enc_lower, enc_name in enc_names_lower.items():
                if (
                    enc_lower == activity_lower
                    or enc_lower in activity_lower
                    or activity_lower in enc_lower
                ):
                    matched_enc = enc_name
                    break
            if matched_enc:
                entry["encounterType"] = matched_enc
                prog = program_encounter_map.get(matched_enc)
                if prog:
                    entry["formType"] = "ProgramEncounter"
                    entry["program"] = prog
                else:
                    entry["formType"] = "Encounter"

        mapping[activity] = entry

    return mapping


# ── Form Sheet Parser ────────────────────────────────────────────────────────


def _match_sheet_to_w3h(
    sheet_name: str, w3h_mapping: dict[str, dict[str, str | None]]
) -> dict[str, str | None] | None:
    """Match a form sheet name to a W3H activity. Picks best scored match."""
    sheet_lower = sheet_name.strip().lower()

    # Exact match
    for activity, entry in w3h_mapping.items():
        if activity.lower() == sheet_lower:
            return entry

    # Scored matching
    sheet_tokens = set(sheet_lower.split())
    best_score = 0.0
    best_entry = None

    for activity, entry in w3h_mapping.items():
        act_lower = activity.lower()

        if act_lower in sheet_lower or sheet_lower in act_lower:
            score = min(len(act_lower), len(sheet_lower)) / max(
                len(act_lower), len(sheet_lower)
            )
            score += 1.0
        else:
            act_tokens = set(act_lower.split())
            if not sheet_tokens or not act_tokens:
                continue
            overlap = len(sheet_tokens & act_tokens)
            score = overlap / max(len(sheet_tokens), len(act_tokens))

        if score > best_score:
            best_score = score
            best_entry = entry

    if best_score >= 0.5:
        return best_entry
    return None


def parse_unified_modelling(
    df: pd.DataFrame,
) -> tuple[
    list[SubjectTypeSpec],
    list[ProgramSpec],
    list[EncounterTypeSpec],
    list[AddressLevelSpec],
]:
    """Parse a unified modelling sheet where Type column identifies entity kind.

    Format:
      Type           | Name         | Subject Type | Color | ...
      Subject Type   | Individual   | Person       |       |
      Subject Type   | Household    | Household    |       |
      Program        | Hypertension |              | #368  |
      Encounter Type | ANC Followup |              |       |
      Address Level  | Village      | Level 1      |       |
    """
    type_col = _find_col(df, "Type")
    name_col = _find_col(df, "Name")
    if not type_col or not name_col:
        return [], [], [], []

    subject_types: list[SubjectTypeSpec] = []
    programs: list[ProgramSpec] = []
    encounters: list[EncounterTypeSpec] = []
    address_levels: list[AddressLevelSpec] = []

    st_col = _find_col(df, "Subject Type", "Entity Type")
    color_col = _find_col(df, "Color", "Colour")

    seen_st: set[str] = set()
    seen_prog: set[str] = set()
    seen_enc: set[str] = set()
    seen_addr: set[str] = set()

    for _, row in df.iterrows():
        row_type = _clean(row.get(type_col, "")).lower()
        name = _clean(row.get(name_col, ""))
        if not name or name.startswith("---"):
            continue

        if row_type == "subject type":
            if name.lower() in seen_st:
                continue
            seen_st.add(name.lower())
            raw_type = _clean(row.get(st_col, "Person")).lower() if st_col else "person"
            avni_type = _SUBJECT_TYPE_MAP.get(raw_type, "Person")
            subject_types.append(
                SubjectTypeSpec(
                    name=name,
                    type=avni_type,
                    allowProfilePicture=False,
                    uniqueName=False,
                    lastNameOptional=True,
                )
            )

        elif row_type == "program":
            if name.lower() in seen_prog:
                continue
            seen_prog.add(name.lower())
            colour = _clean(row.get(color_col, "#4A148C")) if color_col else "#4A148C"
            target = ""
            if st_col:
                target = _clean(row.get(st_col, ""))
            if not target and subject_types:
                target = subject_types[0].name
            programs.append(
                ProgramSpec(
                    name=name,
                    target_subject_type=target,
                    colour=colour or "#4A148C",
                )
            )

        elif row_type == "encounter type":
            if name.lower() in seen_enc:
                continue
            seen_enc.add(name.lower())
            encounters.append(
                EncounterTypeSpec(
                    name=name,
                    program_name="",
                    subject_type="",
                    is_program_encounter=False,
                    is_scheduled=True,
                )
            )

        elif row_type == "address level":
            if name.lower() in seen_addr:
                continue
            seen_addr.add(name.lower())
            # Try to extract level number
            level_str = _clean(row.get(st_col, "")) if st_col else ""
            level = 1
            level_match = re.search(r"(\d+)", level_str)
            if level_match:
                level = int(level_match.group(1))
            address_levels.append(
                AddressLevelSpec(
                    name=name,
                    level=level,
                    parent=None,
                )
            )

    # Fix address level parent chain (highest level = no parent, each lower level → parent is one above)
    if address_levels:
        sorted_levels = sorted(address_levels, key=lambda a: a.level, reverse=True)
        for i in range(1, len(sorted_levels)):
            sorted_levels[i].parent = sorted_levels[i - 1].name

    return subject_types, programs, encounters, address_levels


def _detect_skip_logic_patterns(fields: list[FieldSpec]) -> None:
    """Auto-detect skip logic from field relationships. Mutates fields in place.

    Patterns detected:
    1. "Others" pattern: Coded field with "Others" option + next field "Specify"/"Other" → show when Others
    2. Sub-type pattern: "X Type" (Coded) + "X Sub-Type" (Coded) → show when parent answer matches
    3. Yes/No detail: Coded Yes/No + "X Details"/"X Reason" → show when Yes
    """
    for i, field in enumerate(fields):
        if field.skipLogic:
            continue  # Already has explicit skip logic from SRS

        name_lower = field.name.lower()

        # Pattern 1: "Specify" / "Other details" after a Coded field with "Others" option
        if any(
            kw in name_lower
            for kw in ("specify", "other detail", "if other", "details - ")
        ):
            # Look backwards for the nearest Coded field with "Others" option
            for j in range(i - 1, max(i - 5, -1), -1):
                prev = fields[j]
                if prev.dataType == "Coded" and prev.options:
                    has_others = any(
                        o.lower() in ("others", "other", "na", "not applicable")
                        for o in prev.options
                    )
                    if has_others:
                        other_val = next(
                            o for o in prev.options if o.lower() in ("others", "other")
                        )
                        field.skipLogic = SkipLogicSpec(
                            dependsOn=prev.name,
                            condition="containsAnswerConceptName",
                            value=other_val,
                        )
                        break

        # Pattern 2: Sub-type field following a "Type" field
        # "Health Activity Sub-Type" → parent "Activity Type", qualifier "Health"
        if (
            "sub-type" in name_lower
            or "sub type" in name_lower
            or "subtype" in name_lower
        ):
            # Strip "sub-type" to get the rest, then find the parent by word overlap
            stripped = (
                name_lower.replace("sub-type", "")
                .replace("sub type", "")
                .replace("subtype", "")
                .strip()
            )
            stripped_words = set(re.findall(r"[a-z]+", stripped))
            for j in range(i - 1, max(i - 10, -1), -1):
                prev = fields[j]
                if prev.dataType != "Coded" or not prev.options:
                    continue
                prev_lower = prev.name.lower()
                prev_words = set(re.findall(r"[a-z]+", prev_lower))
                # Parent must share at least one content word (not "type")
                common = (stripped_words & prev_words) - {"type"}
                if common:
                    # Qualifier = words in sub-type name NOT in parent
                    qualifier_words = stripped_words - prev_words - {"type"}
                    qualifier = " ".join(qualifier_words)
                    if qualifier:
                        match_opt = next(
                            (o for o in prev.options if qualifier in o.lower()),
                            None,
                        )
                        if match_opt:
                            field.skipLogic = SkipLogicSpec(
                                dependsOn=prev.name,
                                condition="containsAnswerConceptName",
                                value=match_opt,
                            )
                            break

        # Pattern 3: Yes/No detail — "X Details" / "X Reason" after "X" (Yes/No)
        if field.skipLogic:
            continue
        for suffix in ("details", "reason", "description", "explanation"):
            if suffix in name_lower:
                base = (
                    name_lower.replace(suffix, "")
                    .strip()
                    .rstrip("-")
                    .rstrip("–")
                    .strip()
                )
                if not base:
                    continue
                for j in range(i - 1, max(i - 5, -1), -1):
                    prev = fields[j]
                    if prev.dataType == "Coded" and prev.options:
                        prev_lower = prev.name.lower()
                        if base in prev_lower or prev_lower in base:
                            yes_opts = [
                                o for o in prev.options if o.lower() in ("yes", "true")
                            ]
                            if yes_opts:
                                field.skipLogic = SkipLogicSpec(
                                    dependsOn=prev.name,
                                    condition="containsAnswerConceptName",
                                    value=yes_opts[0],
                                )
                                break
                if field.skipLogic:
                    break


def parse_form_df(
    df: pd.DataFrame,
    sheet_name: str,
    entity_mapping: dict[str, str | None] | None,
    header_offset: int = 0,
) -> FormSpec | None:
    """
    Parse a form DataFrame into a FormSpec.
    header_offset=0: headers in row 0 (standard CSV/clean Excel)
    header_offset=1: headers in row 1 (scoping Excel with category row in row 0)
    """
    if header_offset == 1:
        # Re-read with row 1 as header
        if df.shape[0] < 3:
            return None
        # Use raw positional access — row 0 = categories, row 1 = headers, row 2+ = data
        data_start = 2
    else:
        data_start = 1  # row 0 = headers, row 1+ = data

    # Find columns by fuzzy matching on the header row
    header_row_idx = header_offset  # 0 or 1
    header_row = {
        str(df.iloc[header_row_idx, c]).strip().lower(): c for c in range(df.shape[1])
    }

    def _col_idx(*names: str) -> int | None:
        for n in names:
            n_lower = n.lower()
            for h, idx in header_row.items():
                if n_lower == h or n_lower in h or h in n_lower:
                    return idx
        return None

    page_idx = _col_idx("page name", "page", "section")
    field_idx = _col_idx("field name", "field", "question", "name")
    dtype_idx = _col_idx("data type", "datatype", "type")
    mand_idx = _col_idx("mandatory")
    options_idx = _col_idx("options", "option")
    selection_idx = _col_idx("selection type", "selection")
    min_max_idx = _col_idx("max and min", "min and max", "limit", "range")
    unit_idx = _col_idx("unit")
    skip_idx = _col_idx("when to show", "skip logic", "condition", "visibility")

    if field_idx is None:
        return None

    current_section = "General Information"
    fields: list[FieldSpec] = []

    for row_idx in range(data_start, df.shape[0]):
        row = df.iloc[row_idx]
        field_name = (
            _clean_field_name(row.iloc[field_idx]) if field_idx is not None else ""
        )
        if not field_name:
            continue

        if page_idx is not None:
            page = _clean(row.iloc[page_idx])
            if page:
                current_section = page

        raw_dtype = (
            _clean(row.iloc[dtype_idx]).lower() if dtype_idx is not None else "text"
        )
        avni_dtype = _map_data_type(raw_dtype)

        # Detect dtype column displacement:
        # Case A — dtype cell contains comma/newline-separated values (options bled in):
        #   treat field as Coded and carry those values forward as options candidate.
        _dtype_displaced_options: list[str] | None = None
        if dtype_idx is not None and avni_dtype == "Text" and raw_dtype:
            if "," in raw_dtype or "\n" in raw_dtype or ";" in raw_dtype:
                _dtype_displaced_options = _parse_options(row.iloc[dtype_idx])
                if _dtype_displaced_options:
                    avni_dtype = "Coded"
                    # Try to recover the real type from adjacent columns
                    for scan_idx in range(
                        dtype_idx + 1, min(dtype_idx + 3, row.shape[0])
                    ):
                        candidate_type = _clean(row.iloc[scan_idx]).lower()
                        if candidate_type in _DATA_TYPE_MAP:
                            avni_dtype = _DATA_TYPE_MAP[candidate_type]
                            _dtype_displaced_options = (
                                None  # not displaced options after all
                            )
                            break

        # Case B — dtype cell has an unrecognised single token but an adjacent cell has a
        #   valid type keyword (one-column right-shift of the whole row):
        if (
            dtype_idx is not None
            and avni_dtype == "Text"
            and raw_dtype
            and "," not in raw_dtype
        ):
            for scan_idx in range(dtype_idx + 1, min(dtype_idx + 3, row.shape[0])):
                candidate_type = _clean(row.iloc[scan_idx]).lower()
                if candidate_type in _DATA_TYPE_MAP:
                    avni_dtype = _DATA_TYPE_MAP[candidate_type]
                    break

        mandatory = _parse_yes_no(row.iloc[mand_idx]) if mand_idx is not None else False

        options = None
        _displaced_type_keyword: str | None = None  # type keyword found in options col
        if options_idx is not None:
            raw_options_val = _clean(row.iloc[options_idx]).lower()
            if raw_options_val in _DATA_TYPE_MAP:
                # Column displacement detected: the options cell contains a type keyword.
                # Record it for selection_type inference and scan adjacent columns for
                # the actual options.
                _displaced_type_keyword = raw_options_val
                for scan_idx in range(
                    options_idx + 1, min(options_idx + 4, row.shape[0])
                ):
                    candidate = _parse_options(row.iloc[scan_idx])
                    if candidate:
                        options = candidate
                        break
            else:
                options = _parse_options(row.iloc[options_idx]) or None

        # Use dtype-displaced options as fallback when options column had nothing
        if not options and _dtype_displaced_options:
            options = _dtype_displaced_options

        # If boolean/yes-no type detected, auto-add options
        if (
            avni_dtype == "Coded"
            and not options
            and raw_dtype in ("boolean", "bool", "yes/no")
        ):
            options = ["Yes", "No"]

        selection_type = None
        if selection_idx is not None:
            sel_raw = _clean(row.iloc[selection_idx]).lower()
            if "multi" in sel_raw:
                selection_type = "MultiSelect"
            elif "single" in sel_raw:
                selection_type = "SingleSelect"
        # Infer selection type from displaced keyword when no explicit selection column
        if selection_type is None and _displaced_type_keyword:
            if "multi" in _displaced_type_keyword:
                selection_type = "MultiSelect"
            elif "single" in _displaced_type_keyword:
                selection_type = "SingleSelect"

        min_val, max_val = None, None
        if min_max_idx is not None:
            min_val, max_val = _parse_min_max(row.iloc[min_max_idx])

        unit = _clean(row.iloc[unit_idx]) if unit_idx is not None else None
        if not unit:
            unit = None

        skip_logic = None
        if skip_idx is not None:
            when_to_show = _clean(row.iloc[skip_idx])
            if when_to_show:
                skip_logic = SkipLogicSpec(
                    dependsOn=when_to_show, condition="raw", value=when_to_show
                )

        fields.append(
            FieldSpec(
                name=field_name,
                dataType=avni_dtype,
                mandatory=mandatory,
                group=current_section,
                unit=unit,
                min=min_val,
                max=max_val,
                options=options,
                selectionType=selection_type,
                skipLogic=skip_logic,
            )
        )

    if not fields:
        return None

    # Auto-detect skip logic patterns from field relationships
    _detect_skip_logic_patterns(fields)

    form_type = "Encounter"
    subject_type = None
    program = None
    encounter_type = None

    if entity_mapping:
        form_type = entity_mapping.get("formType") or "Encounter"
        subject_type = entity_mapping.get("subjectType")
        program = entity_mapping.get("program")
        encounter_type = entity_mapping.get("encounterType")

    # Build sections
    sections: list[SectionSpec] = []
    cur_name: str | None = None
    cur_fields: list[FieldSpec] = []
    for f in fields:
        sec_name = f.group or "General Information"
        if sec_name != cur_name:
            if cur_fields:
                sections.append(
                    SectionSpec(
                        name=cur_name or "General Information", fields=cur_fields
                    )
                )
            cur_name = sec_name
            cur_fields = [f]
        else:
            cur_fields.append(f)
    if cur_fields:
        sections.append(
            SectionSpec(name=cur_name or "General Information", fields=cur_fields)
        )

    return FormSpec(
        name=sheet_name.strip(),
        formType=form_type,
        subjectType=subject_type,
        program=program,
        encounterType=encounter_type,
        sections=sections,
    )


# ── File Loaders ─────────────────────────────────────────────────────────────


def _load_xlsx(path: Path) -> list[tuple[str, pd.DataFrame]]:
    """Load all sheets from an Excel file. Returns [(sheet_name, df), ...].

    Strikethrough cells are blanked: pandas' read_excel discards formatting,
    so an openpyxl side-pass identifies them and nulls the corresponding
    DataFrame entries. Convention: a struck-through field is out of scope.
    """
    xf = pd.ExcelFile(path)
    sheets: list[tuple[str, pd.DataFrame]] = []
    struck_by_sheet: dict[str, set[tuple[int, int]]] = (
        _struck_cell_coords_per_sheet(path)
        if path.suffix.lower() == ".xlsx"
        else {}
    )
    for name in xf.sheet_names:
        try:
            df = pd.read_excel(xf, sheet_name=name, header=None)
            if df.empty:
                continue
            struck = struck_by_sheet.get(name)
            if struck:
                for (r, c) in struck:
                    if r < df.shape[0] and c < df.shape[1]:
                        df.iat[r, c] = pd.NA
            sheets.append((name, df))
        except Exception as e:
            logger.warning("Failed to read sheet '%s' in %s: %s", name, path, e)
    return sheets


def _struck_cell_coords_per_sheet(path: Path) -> dict[str, set[tuple[int, int]]]:
    """Return `{sheet_name: {(row_idx, col_idx), …}}` for strikethrough cells.

    Indices are 0-based (openpyxl is 1-based — offset applied below).
    Best-effort: workbook-open failures return an empty dict.
    """
    import openpyxl

    out: dict[str, set[tuple[int, int]]] = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("openpyxl could not open %s for strikethrough scan: %s", path, exc)
        return out
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            struck: set[tuple[int, int]] = set()
            for row in ws.iter_rows():
                for cell in row:
                    font = getattr(cell, "font", None)
                    if font is not None and getattr(font, "strike", False):
                        struck.add((cell.row - 1, cell.column - 1))
            if struck:
                out[sheet_name] = struck
    finally:
        wb.close()
    return out


def _load_file(path: Path) -> list[tuple[str, pd.DataFrame]]:
    """Load any supported file format."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        return _load_xlsx(path)
    logger.warning("Unsupported file type: %s", path)
    return []


# ── Post-processing ──────────────────────────────────────────────────────────


def _resolve_form_subject_types(
    forms: list[FormSpec],
    encounter_types: list[EncounterTypeSpec],
    subject_types: list[SubjectTypeSpec],
    programs: list[ProgramSpec],
) -> None:
    """Fix formType, encounterType, subjectType, and program on forms. Mutates in place."""
    enc_to_subject: dict[str, str] = {}
    enc_to_program: dict[str, str] = {}
    for et in encounter_types:
        if et.subject_type:
            enc_to_subject[et.name.lower()] = et.subject_type
        if et.program_name:
            enc_to_program[et.name.lower()] = et.program_name

    prog_to_subject = {
        p.name.lower(): p.target_subject_type for p in programs if p.target_subject_type
    }

    st_names = {st.name for st in subject_types}
    prog_names = {p.name for p in programs}
    enc_names = {et.name for et in encounter_types}

    for form in forms:
        name_lower = form.name.lower()

        # Step 0: Fix formType based on form name patterns
        if form.formType == "Encounter":
            if any(kw in name_lower for kw in ("registration", "profile", "details")):
                base = name_lower
                for kw in ("registration", "profile", "details"):
                    base = base.replace(kw, "").strip()
                matched_st = _fuzzy_match(base, st_names)
                if matched_st:
                    form.formType = "IndividualProfile"
                    form.subjectType = matched_st
                elif len(subject_types) == 1:
                    form.formType = "IndividualProfile"
                    form.subjectType = subject_types[0].name
            elif any(kw in name_lower for kw in ("enrolment", "enrollment")):
                matched_prog = _fuzzy_match(form.name, prog_names)
                if matched_prog:
                    prog = next(p for p in programs if p.name == matched_prog)
                    form.formType = "ProgramEnrolment"
                    form.program = prog.name
                    form.subjectType = prog.target_subject_type
            elif "exit" in name_lower:
                form_base = name_lower.replace("exit", "").strip().rstrip("-").strip()
                prog_bases = {
                    p.name: p.name.lower()
                    .replace("enrollment", "")
                    .replace("enrolment", "")
                    .strip()
                    for p in programs
                }
                matched_prog = _fuzzy_match(form_base, set(prog_bases.values()))
                if matched_prog:
                    prog = next(p for p, b in prog_bases.items() if b == matched_prog)
                    real_prog = next(p for p in programs if p.name == prog)
                    form.formType = "ProgramExit"
                    form.program = real_prog.name
                    form.subjectType = real_prog.target_subject_type

        # Step 1: Match form to encounter type by name (fuzzy)
        if not form.encounterType and form.formType in (
            "Encounter",
            "ProgramEncounter",
            "IndividualEncounterCancellation",
            "ProgramEncounterCancellation",
        ):
            match_name = (
                form.name.replace(" Cancellation", "")
                .replace(" cancellation", "")
                .strip()
            )
            matched_enc = _fuzzy_match(match_name, enc_names)
            if matched_enc:
                et = next(e for e in encounter_types if e.name == matched_enc)
                form.encounterType = et.name
                if et.is_program_encounter and et.program_name:
                    form.formType = (
                        "ProgramEncounter"
                        if "cancellation" not in name_lower
                        else "ProgramEncounterCancellation"
                    )
                    form.program = et.program_name

        # Step 2: Resolve subjectType from encounterType
        if not form.subjectType and form.encounterType:
            subject = enc_to_subject.get(form.encounterType.lower())
            if subject:
                form.subjectType = subject

        # Step 3: Resolve subjectType from program (fuzzy)
        if not form.subjectType and form.program:
            matched_prog = _fuzzy_match(form.program, set(prog_to_subject.keys()))
            if matched_prog:
                form.subjectType = prog_to_subject[matched_prog]

        # Step 4: Resolve subjectType for registration forms (fuzzy)
        if not form.subjectType and form.formType == "IndividualProfile":
            matched_st = _fuzzy_match(form.name, st_names)
            if matched_st:
                form.subjectType = matched_st

        # Step 5: Resolve program from encounterType
        if not form.program and form.encounterType:
            prog = enc_to_program.get(form.encounterType.lower())
            if prog:
                form.program = prog

        # Step 6: Fallback — single subject type
        if not form.subjectType and len(subject_types) == 1:
            form.subjectType = subject_types[0].name


# ── Main Entry Point ─────────────────────────────────────────────────────────


def parse_scoping_docs(
    file_paths: list[str | Path],
) -> tuple["EntitySpec", list[dict]]:
    """
    Parse one or more SRS/scoping/modelling .xlsx files and consolidate into
    a validated EntitySpec.

    Each sheet is auto-classified by its content and parsed accordingly.
    All results are merged with deduplication.

    Returns:
        (EntitySpec, misc_sheets) where misc_sheets is a list of dicts for
        sheets that could not be classified, each containing:
        {name, file, columns, rows (first 20 non-empty rows as list-of-dicts)}
    """
    all_address: list[AddressLevelSpec] = []
    all_subjects: list[SubjectTypeSpec] = []
    all_programs: list[ProgramSpec] = []
    all_encounters: list[EncounterTypeSpec] = []
    all_forms: list[FormSpec] = []
    w3h_dfs: list[pd.DataFrame] = []
    form_sheets: list[tuple[str, pd.DataFrame, int]] = []  # (name, df, header_offset)
    misc_sheets: list[dict] = []
    # Per-form rule intents harvested from a dedicated "Rules" / "Form Rules"
    # tab. Each row in that tab corresponds to one form (cancellation and exit
    # forms get their own rows, distinct from the parent encounter type).
    # Stitched onto FormSpec.rule_intents at the end of phase 5.
    intents_by_form_name: dict[str, dict[str, str]] = {}

    seen: dict[str, set[str]] = {
        "address": set(),
        "subject": set(),
        "program": set(),
        "encounter": set(),
        "form": set(),
    }

    # Phase 1: Load all files and classify each sheet
    for file_path in file_paths:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        logger.info("Loading file: %s", path)
        sheets = _load_file(path)

        for sheet_name, df in sheets:
            if df.empty or df.shape[1] < 2:
                continue

            # Try classifying with row 0 as headers
            df_with_header = df.copy()
            df_with_header.columns = [
                str(df.iloc[0, c]).strip() if df.shape[0] > 0 else f"col{c}"
                for c in range(df.shape[1])
            ]
            df_with_header = df_with_header.iloc[1:].reset_index(drop=True)
            df_with_header = df_with_header.dropna(how="all")

            classification = _classify_sheet(df_with_header, sheet_name)

            # If unknown, try with row 1 as headers (scoping doc pattern)
            if classification == "unknown" and df.shape[0] >= 3:
                df_alt = df.copy()
                df_alt.columns = [
                    str(df.iloc[1, c]).strip() if df.shape[0] > 1 else f"col{c}"
                    for c in range(df.shape[1])
                ]
                alt_class = _classify_sheet(df_alt, sheet_name)
                if alt_class in ("form", "form_offset1"):
                    classification = "form_offset1"
                elif alt_class != "unknown":
                    classification = alt_class
                    df_with_header = (
                        df_alt.iloc[2:].reset_index(drop=True).dropna(how="all")
                    )

            if classification == "form_offset1":
                classification = "form"

            logger.info("  Sheet '%s': classified as %s", sheet_name, classification)

            subject_type_names = {st.name for st in all_subjects}
            program_names = {p.name for p in all_programs}

            if classification == "unified_modelling":
                u_sts, u_progs, u_encs, u_addrs = parse_unified_modelling(
                    df_with_header
                )
                for st in u_sts:
                    if st.name.lower() not in seen["subject"]:
                        all_subjects.append(st)
                        seen["subject"].add(st.name.lower())
                for p in u_progs:
                    if p.name.lower() not in seen["program"]:
                        all_programs.append(p)
                        seen["program"].add(p.name.lower())
                for et in u_encs:
                    if et.name.lower() not in seen["encounter"]:
                        all_encounters.append(et)
                        seen["encounter"].add(et.name.lower())
                for al in u_addrs:
                    if al.name.lower() not in seen["address"]:
                        all_address.append(al)
                        seen["address"].add(al.name.lower())

            elif classification == "location":
                for al in parse_location_hierarchy(df_with_header):
                    if al.name.lower() not in seen["address"]:
                        all_address.append(al)
                        seen["address"].add(al.name.lower())

            elif classification == "subject_type":
                for st in parse_subject_types(df_with_header):
                    if st.name.lower() not in seen["subject"]:
                        all_subjects.append(st)
                        seen["subject"].add(st.name.lower())

            elif classification == "program":
                for prog in parse_programs(df_with_header, subject_type_names):
                    if prog.name.lower() not in seen["program"]:
                        all_programs.append(prog)
                        seen["program"].add(prog.name.lower())

            elif classification in ("encounter", "program_encounter"):
                is_prog = classification == "program_encounter"
                for et in parse_encounters(
                    df_with_header, subject_type_names, program_names, is_prog
                ):
                    if et.name.lower() not in seen["encounter"]:
                        all_encounters.append(et)
                        seen["encounter"].add(et.name.lower())

            elif classification == "rules":
                name_col = find_form_name_column(df_with_header)
                if name_col:
                    intents_by_form_name.update(
                        extract_rule_intents(df_with_header, name_col)
                    )

            elif classification == "w3h":
                w3h_dfs.append(df_with_header)

            elif classification == "form":
                # Auto-detect header row: if row 0 has form headers → offset=0
                # Otherwise (scoping format with title in row 0) → offset=1
                offset = 0
                if df.shape[0] > 0:
                    row0_vals = [
                        str(df.iloc[0, c]).strip().lower()
                        for c in range(min(df.shape[1], 6))
                    ]
                    has_form_headers = any(
                        h in row0_vals
                        for h in (
                            "field name",
                            "data type",
                            "page name",
                            "field",
                            "mandatory",
                        )
                    )
                    offset = 0 if has_form_headers else 1
                form_sheets.append((sheet_name, df, offset))

            else:
                # Unknown — capture for agent inspection
                cols = [str(c) for c in df_with_header.columns.tolist()]
                rows = (
                    df_with_header.head(3)
                    .dropna(how="all")
                    .fillna("")
                    .infer_objects(copy=False)
                    .to_dict(orient="records")
                )
                misc_sheets.append(
                    {
                        "name": sheet_name,
                        "file": path.name,
                        "columns": cols,
                        "rows": rows,
                    }
                )

    # Default address levels if none found
    if not all_address:
        logger.warning(
            "No location hierarchy found — using default State/District/Village"
        )
        all_address = [
            AddressLevelSpec(name="State", level=3, parent=None),
            AddressLevelSpec(name="District", level=2, parent="State"),
            AddressLevelSpec(name="Village", level=1, parent="District"),
        ]

    # Phase 2: Build consolidated context
    subject_type_names = {st.name for st in all_subjects}
    encounter_type_names = {et.name for et in all_encounters}
    program_encounter_map: dict[str, str] = {
        et.name: et.program_name for et in all_encounters if et.program_name
    }

    # Phase 3: Parse W3H for form-entity mapping
    w3h_mapping: dict[str, dict[str, str | None]] = {}
    for w3h_df in w3h_dfs:
        w3h_mapping.update(
            parse_w3h(
                w3h_df, subject_type_names, encounter_type_names, program_encounter_map
            )
        )
    logger.info("W3H mapping: %d activities", len(w3h_mapping))

    # Phase 4: Parse form sheets using consolidated context
    for sheet_name, df, offset in form_sheets:
        if sheet_name.lower() in seen["form"]:
            continue

        entity_mapping = _match_sheet_to_w3h(sheet_name, w3h_mapping)
        form = parse_form_df(df, sheet_name, entity_mapping, header_offset=offset)
        if form:
            all_forms.append(form)
            seen["form"].add(sheet_name.lower())

    # Phase 5a: Resolve missing subject_type on encounters via programs
    prog_to_subject: dict[str, str] = {}
    for prog in all_programs:
        if prog.target_subject_type:
            prog_to_subject[prog.name.lower()] = prog.target_subject_type
    prog_names_for_match = set(prog_to_subject.keys())
    st_names_set = {st.name for st in all_subjects}
    for et in all_encounters:
        if not et.subject_type and et.program_name:
            # Try matching program_name to actual programs
            matched = _fuzzy_match(et.program_name, prog_names_for_match)
            resolved = prog_to_subject.get(matched) if matched else None
            if resolved:
                et.subject_type = resolved
            else:
                # program_name might actually be a subject type name
                # (some SRS docs put subject type in the program column)
                matched_st = _fuzzy_match(et.program_name, st_names_set)
                if matched_st:
                    et.subject_type = matched_st
                    et.program_name = ""  # Clear — it's a subject type, not a program
        # If still no subject_type and only one subject type exists, use it
        if not et.subject_type and len(all_subjects) == 1:
            et.subject_type = all_subjects[0].name

    # Phase 5b: Resolve missing subjectType on forms
    _resolve_form_subject_types(all_forms, all_encounters, all_subjects, all_programs)

    # Phase 5c: Attach rule intents harvested from the Rules tab
    if intents_by_form_name:
        apply_intents_to_forms(all_forms, intents_by_form_name)
        attached = sum(1 for f in all_forms if f.rule_intents)
        logger.info(
            "Attached rule intents to %d of %d form(s)",
            attached, len(all_forms),
        )

    # Phase 6: Always include default group
    groups = [GroupSpec(name="Everyone", has_all_privileges=True)]

    logger.info(
        "Consolidated: %d addresses, %d subjects, %d programs, "
        "%d encounters, %d forms, %d groups from %d file(s)",
        len(all_address),
        len(all_subjects),
        len(all_programs),
        len(all_encounters),
        len(all_forms),
        len(groups),
        len(file_paths),
    )

    return EntitySpec(
        subject_types=all_subjects,
        programs=all_programs,
        encounter_types=all_encounters,
        address_levels=all_address,
        groups=groups,
        forms=all_forms,
    ), misc_sheets

